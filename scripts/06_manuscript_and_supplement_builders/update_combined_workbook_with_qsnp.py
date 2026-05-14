from pathlib import Path
import shutil

import pandas as pd
from openpyxl import load_workbook


base_dir = Path(r"D:\metabolic\GWAS\genomicgem_main_zgt4_nonproportion\supplement_combined_lipid_nonlipid_final8")
src_wb = base_dir / "metabolic_lipid_nonlipid_final8_combined_supplement_workbook_with_external_validation_and_neff_formal_fit_revised.xlsx"
dst_wb = base_dir / "metabolic_lipid_nonlipid_final8_combined_supplement_workbook_with_external_validation_and_neff_formal_fit_qsnp_revised.xlsx"

qdir = Path(r"D:\metabolic\GWAS\genomicgem_main_zgt4_nonproportion\q_snp_ld_clump_final8")
figdir = Path(r"D:\metabolic\GWAS\genomicgem_main_zgt4_nonproportion\q_snp_figures_final8")

files = {
    "33_QSNP_LDClump_Summary": qdir / "final8_qsnp_ldclump_summary.tsv",
    "34_FactorLeadLoci_LDClump": qdir / "final8_factor_lead_loci_ldclump.tsv",
    "35_QSNP_LeadLoci_LDClump": qdir / "final8_qsnp_lead_loci_ldclump.tsv",
    "36_Factor_QSNP_LD_Overlap": qdir / "final8_factor_qsnp_ld_overlap_pairs.tsv",
    "37_UniqueFactorLeads_ExclQSNP": qdir / "final8_unique_factor_leads_excluding_qsnp_ld.tsv",
    "38_QSNP_FigureManifest": figdir / "q_snp_figure_manifest.tsv",
}

descriptions = {
    "33_QSNP_LDClump_Summary": "Reference-based LD-clumping summary for factor hits and Q_SNP hits, including exact lead-SNP overlap and unique factor lead loci after excluding LD with Q_SNP lead loci.",
    "34_FactorLeadLoci_LDClump": "LD-clumped lead loci for factor GWAS results using g1000_eur, P < 5e-8, r2 < 0.1, and 1000 kb windows.",
    "35_QSNP_LeadLoci_LDClump": "LD-clumped lead loci for Q_SNP results using g1000_eur, P < 5e-8, r2 < 0.1, and 1000 kb windows.",
    "36_Factor_QSNP_LD_Overlap": "Pairwise LD overlap records between factor lead SNPs and Q_SNP lead SNPs at r2 >= 0.1 within 1000 kb.",
    "37_UniqueFactorLeads_ExclQSNP": "Factor lead loci retained after excluding loci in LD with Q_SNP lead loci.",
    "38_QSNP_FigureManifest": "Manifest of Q_SNP Manhattan/QQ figures generated for lipid and nonlipid modules.",
}

notes = {
    "33_QSNP_LDClump_Summary": "Lead loci were defined by PLINK clumping against D:/SMR/g1000/g1000_eur with genome-wide significance threshold P < 5e-8, r2 < 0.1, and 1000 kb windows.",
    "34_FactorLeadLoci_LDClump": "These loci correspond to genome-wide significant factor hits after reference-based LD pruning.",
    "35_QSNP_LeadLoci_LDClump": "These loci correspond to genome-wide significant heterogeneity hits after reference-based LD pruning.",
    "36_Factor_QSNP_LD_Overlap": "This sheet can be used to support statements about overlap between factor-discovery loci and heterogeneity loci.",
    "37_UniqueFactorLeads_ExclQSNP": "This sheet supports statements of the form 'unique factor hits excluding LD with Q_SNP hits'.",
    "38_QSNP_FigureManifest": "Figure files are stored under D:/metabolic/GWAS/genomicgem_main_zgt4_nonproportion/q_snp_figures_final8.",
}


def write_df(ws, df):
    ws.delete_rows(1, ws.max_row)
    rows = [df.columns.tolist()] + df.astype(object).where(pd.notnull(df), "").values.tolist()
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)


shutil.copy2(src_wb, dst_wb)
wb = load_workbook(dst_wb)

for sheet_name, path in files.items():
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    df = pd.read_csv(path, sep="\t")
    write_df(ws, df)

index_ws = wb["00_Index"]
for sheet_name in files:
    index_ws.append([sheet_name, descriptions[sheet_name], "", notes[sheet_name]])

index_ws["F4"] = "This workbook extends the formal-fit revised workbook by appending LD-clumped Q_SNP core results and figure manifests."
wb.save(dst_wb)
print(dst_wb)

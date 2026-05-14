from __future__ import annotations

import copy
import csv
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


ROOT = Path(r"D:\codex\GenomicSEM\metabolic")
COMBINED = Path(
    r"D:\metabolic\GWAS\genomicgem_main_zgt4_nonproportion"
    r"\supplement_combined_lipid_nonlipid_final8"
    r"\metabolic_lipid_nonlipid_final8_combined_supplement_workbook_with_external_validation_and_neff_formal_fit_qsnp_revised.xlsx"
)
V5 = ROOT / (
    r"postgwas_ad_pdlbd\results\22_supplement_tables_lipid8_F2_AD"
    r"\metabolic_factor_triplet_supplementary_tables_v5_slim_submission_with_target_validation.xlsx"
)
NDD_LDSC = ROOT / (
    r"postgwas_ad_pdlbd\results\03_ldsc_metabolic_factors_vs_ndd"
    r"\metabolic_factors_vs_ndd_requested_pairs.tsv"
)
OUT = ROOT / (
    r"postgwas_ad_pdlbd\results\22_supplement_tables_lipid8_F2_AD"
    r"\metabolic_factor_triplet_supplementary_tables_v10_full_validation_submission.xlsx"
)
LIPID_INTERNAL_DIR = Path(
    r"D:\metabolic\GWAS\genomicgem_main_zgt4_nonproportion\step24_internal_bivariate_lipid_final8"
)
NONLIPID_INTERNAL_DIR = Path(
    r"D:\metabolic\GWAS\genomicgem_main_zgt4_nonproportion\step25_internal_bivariate_nonlipid_final8"
)
LOO_TSV = ROOT / (
    r"postgwas_ad_pdlbd\results\22_supplement_tables_lipid8_F2_AD"
    r"\loo_sensitivity\final8_loo_sensitivity_summary.tsv"
)
MIXER_SUMMARIES = [
    ROOT / r"postgwas_ad_pdlbd\results\05_mixer_lipid8_F2_AD\mixer_lipid8_F2_AD_summary.tsv",
    ROOT / r"postgwas_ad_pdlbd\results\05_mixer_nonlipid8_F1_PD\mixer_nonlipid8_F1_PD_summary.tsv",
    ROOT / r"postgwas_ad_pdlbd\results\05_mixer_lipid8_F1_PD\mixer_lipid8_F1_PD_summary.tsv",
]


HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
TITLE_FONT = Font(color="FFFFFF", bold=True)


def clone_sheet(src_ws, dst_wb: Workbook, title: str):
    ws = dst_wb.create_sheet(title=title)
    for row in src_ws.iter_rows():
        for cell in row:
            new = ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new.font = copy.copy(cell.font)
                new.fill = copy.copy(cell.fill)
                new.border = copy.copy(cell.border)
                new.alignment = copy.copy(cell.alignment)
                new.number_format = cell.number_format
                new.protection = copy.copy(cell.protection)
    for col_letter, dim in src_ws.column_dimensions.items():
        ws.column_dimensions[col_letter].width = dim.width
    for ridx, dim in src_ws.row_dimensions.items():
        ws.row_dimensions[ridx].height = dim.height
    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
    return ws


def read_ws_records(ws):
    rows = list(ws.iter_rows(values_only=True))
    header = [str(x) if x is not None else "" for x in rows[0]]
    records = []
    for r in rows[1:]:
        records.append({header[i]: r[i] if i < len(r) else None for i in range(len(header))})
    return header, records


def write_records(wb: Workbook, title: str, header, records):
    ws = wb.create_sheet(title=title)
    ws.append(header)
    for rec in records:
        ws.append([rec.get(h) for h in header])
    style_sheet(ws)
    return ws


def read_tsv_records(path: Path):
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f, delimiter="\t")
        return reader.fieldnames or [], list(reader)


def write_tsv_sheet(wb: Workbook, title: str, path: Path, extra=None):
    header, records = read_tsv_records(path)
    if extra:
        for rec in records:
            rec.update(extra)
        header = list(header) + [k for k in extra if k not in header]
    return write_records(wb, title, header, records)


def style_sheet(ws):
    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
    for col_idx in range(1, min(ws.max_column, 40) + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter][: min(ws.max_row, 200)]:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 60))
        ws.column_dimensions[letter].width = max(12, min(max_len + 2, 45))


def write_matrix_from_long(comb_wb, out_wb, source_sheet: str, value_col: str, title: str):
    _, records = read_ws_records(comb_wb[source_sheet])
    grouped = {}
    col_traits = {}
    for rec in records:
        module = rec.get("module")
        row_trait = rec.get("row_trait")
        col_trait = rec.get("col_trait")
        value = rec.get(value_col)
        grouped.setdefault(module, {})
        grouped[module].setdefault(row_trait, {})
        grouped[module][row_trait][col_trait] = value
        col_traits.setdefault(module, [])
        if col_trait not in col_traits[module]:
            col_traits[module].append(col_trait)

    output = []
    for module in ["lipid", "nonlipid"]:
        if module not in grouped:
            continue
        output.append({"module": module.upper()})
        header = ["module", "row_trait"] + col_traits[module]
        output.append(dict(zip(header, header)))
        ordered_cols = col_traits[module]
        for row_trait in sorted(grouped[module], key=lambda x: str(x)):
            row_label = row_trait
            if isinstance(row_trait, int) and 1 <= row_trait <= len(ordered_cols):
                row_label = ordered_cols[row_trait - 1]
            elif isinstance(row_trait, str) and str(row_trait).isdigit():
                row_num = int(row_trait)
                if 1 <= row_num <= len(ordered_cols):
                    row_label = ordered_cols[row_num - 1]
            rec = {"module": module, "row_trait": row_label}
            for col_trait in col_traits[module]:
                rec[col_trait] = grouped[module][row_trait].get(col_trait)
            output.append(rec)
        output.append({})

    # convert block-style output into sheet rows
    ws = out_wb.create_sheet(title=title)
    started = False
    current_header = None
    row_idx = 1
    for rec in output:
        if rec == {}:
            row_idx += 1
            continue
        if list(rec.keys()) == ["module"]:
            ws.cell(row=row_idx, column=1, value=rec["module"])
            ws.cell(row=row_idx, column=1).fill = TITLE_FILL
            ws.cell(row=row_idx, column=1).font = TITLE_FONT
            row_idx += 1
            started = False
            continue
        if not started and rec.get("module") == "module":
            current_header = list(rec.keys())
            for c, h in enumerate(current_header, start=1):
                ws.cell(row=row_idx, column=c, value=h)
            started = True
            row_idx += 1
            continue
        for c, h in enumerate(current_header, start=1):
            ws.cell(row=row_idx, column=c, value=rec.get(h))
        row_idx += 1
    style_sheet(ws)
    return ws


def merge_final16(comb_wb, out_wb):
    _, manifest = read_ws_records(comb_wb["14_FinalManifest_Combined"])
    _, ldsc = read_ws_records(comb_wb["15_FinalLDSC_Combined"])
    ldsc_by_key = {}
    for rec in ldsc:
        key = (rec.get("module"), rec.get("trait_code") or rec.get("trait"))
        ldsc_by_key[key] = rec

    header = [
        "module",
        "trait_code",
        "biomarker_name",
        "biomarker_group",
        "final_factor_name",
        "factor_membership_note",
        "h2_observed",
        "h2_se",
        "h2_z",
        "intercept",
        "source_manifest_sheet",
        "source_ldsc_sheet",
    ]
    records = []
    for rec in manifest:
        if not rec.get("module"):
            continue
        key = (rec.get("module"), rec.get("trait_code"))
        lrec = ldsc_by_key.get(key, {})
        records.append(
            {
                "module": rec.get("module"),
                "trait_code": rec.get("trait_code"),
                "biomarker_name": rec.get("biomarker_name") or rec.get("trait_label"),
                "biomarker_group": rec.get("biomarker_group") or rec.get("group"),
                "final_factor_name": rec.get("final_factor_name") or rec.get("factor"),
                "factor_membership_note": rec.get("factor_membership_note") or rec.get("notes"),
                "h2_observed": lrec.get("h2") or lrec.get("h2_observed"),
                "h2_se": lrec.get("h2_se"),
                "h2_z": lrec.get("h2_z"),
                "intercept": lrec.get("intercept"),
                "source_manifest_sheet": "14_FinalManifest_Combined",
                "source_ldsc_sheet": "15_FinalLDSC_Combined",
            }
        )
    write_records(out_wb, "S02_Final16_details", header, records)


def merge_factor_gwas_neff(comb_wb, out_wb, title: str):
    _, files = read_ws_records(comb_wb["21_FactorGWAS_Files"])
    _, neff = read_ws_records(comb_wb["30_FactorGWAS_Neff"])
    neff_by_standard_file = {rec.get("standard_file"): rec for rec in neff}
    all_keys = []
    for rec in files + neff:
        for k in rec.keys():
            if k not in all_keys:
                all_keys.append(k)
    header = ["factor"] + [k for k in all_keys if k != "factor"]
    records = []
    for rec in files:
        factor_id = f"{rec.get('module')}8_{rec.get('factor')}"
        merged = dict(rec)
        merged["factor_id"] = factor_id
        merged.update({k: v for k, v in neff_by_standard_file.get(rec.get("standard_file"), {}).items() if k != "standard_file"})
        records.append(merged)
    write_records(out_wb, title, header, records)


def write_ndd_ldsc(out_wb, title: str):
    with NDD_LDSC.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f, delimiter="\t")
        rows = list(reader)
        header = reader.fieldnames or []
    for rec in rows:
        try:
            rec["fdr_significant_0.05"] = float(rec.get("fdr_rg", "nan")) < 0.05
        except ValueError:
            rec["fdr_significant_0.05"] = ""
        rec["analysis_scope"] = "six metabolic factor GWAS tested against AD, PD, and LBD"
    write_records(
        out_wb,
        title,
        header + ["fdr_significant_0.05", "analysis_scope"],
        rows,
    )


def write_internal_bivar(out_wb, summary_title: str, pair_title: str, manifest_title: str):
    summary_header = None
    summary_records = []
    for module, path in [
        ("lipid", LIPID_INTERNAL_DIR / "lipid_internal_bivariate_ldsc_pairset_summary.tsv"),
        ("nonlipid", NONLIPID_INTERNAL_DIR / "nonlipid_internal_bivariate_ldsc_pairset_summary.tsv"),
    ]:
        header, records = read_tsv_records(path)
        if summary_header is None:
            summary_header = ["module"] + header
        for rec in records:
            rec["module"] = module
            summary_records.append(rec)
    write_records(out_wb, summary_title, summary_header, summary_records)

    pair_header = None
    pair_records = []
    for module, path in [
        ("lipid", LIPID_INTERNAL_DIR / "lipid_internal_bivariate_ldsc_requested_pairs.tsv"),
        ("nonlipid", NONLIPID_INTERNAL_DIR / "nonlipid_internal_bivariate_ldsc_requested_pairs.tsv"),
    ]:
        header, records = read_tsv_records(path)
        if pair_header is None:
            pair_header = ["module"] + header
        for rec in records:
            rec["module"] = module
            pair_records.append(rec)
    write_records(out_wb, pair_title, pair_header, pair_records)

    manifest_header = None
    manifest_records = []
    for module, path in [
        ("lipid", LIPID_INTERNAL_DIR / "lipid_internal_validation_traits_manifest.tsv"),
        ("nonlipid", NONLIPID_INTERNAL_DIR / "nonlipid_internal_validation_traits_manifest.tsv"),
    ]:
        header, records = read_tsv_records(path)
        if manifest_header is None:
            manifest_header = ["module"] + header
        for rec in records:
            rec["module"] = module
            manifest_records.append(rec)
    write_records(out_wb, manifest_title, manifest_header, manifest_records)


def write_internal_bivar_compact(out_wb):
    compact_header = [
        "module",
        "factor",
        "target_trait",
        "biomarker_name",
        "group",
        "pair_set",
        "expected_alignment",
        "rg",
        "rg_se",
        "p_rg",
        "abs_rg",
    ]
    compact_records = []
    for module, path in [
        ("lipid", LIPID_INTERNAL_DIR / "lipid_internal_bivariate_ldsc_requested_pairs.tsv"),
        ("nonlipid", NONLIPID_INTERNAL_DIR / "nonlipid_internal_bivariate_ldsc_requested_pairs.tsv"),
    ]:
        _, records = read_tsv_records(path)
        for rec in records:
            if rec.get("pair_set") not in {"final8_primary", "same_domain_support"}:
                continue
            compact_records.append(
                {
                    "module": module,
                    "factor": rec.get("factor"),
                    "target_trait": rec.get("target_trait"),
                    "biomarker_name": rec.get("biomarker_name"),
                    "group": rec.get("group"),
                    "pair_set": rec.get("pair_set"),
                    "expected_alignment": rec.get("expected_alignment"),
                    "rg": rec.get("rg"),
                    "rg_se": rec.get("rg_se"),
                    "p_rg": rec.get("p_rg"),
                    "abs_rg": rec.get("abs_rg"),
                }
            )
    compact_records.sort(key=lambda r: (r["module"], r["factor"], r["pair_set"], r["target_trait"]))
    write_records(out_wb, "S12_IntFactorMetab_compact", compact_header, compact_records)


def write_factor_trait_map(out_wb):
    header = [
        "module",
        "factor",
        "factor_label",
        "role",
        "trait_code",
        "biomarker_name",
        "group",
        "used_in_factor_model",
        "used_only_for_internal_validation",
    ]
    records = []
    for module, path in [
        ("lipid", LIPID_INTERNAL_DIR / "lipid_internal_validation_traits_manifest.tsv"),
        ("nonlipid", NONLIPID_INTERNAL_DIR / "nonlipid_internal_validation_traits_manifest.tsv"),
    ]:
        _, rows = read_tsv_records(path)
        for rec in rows:
            category = rec.get("category")
            if category == "factor":
                continue
            role = "final8_indicator" if rec.get("subset") == "final8" else "same_domain_support"
            records.append(
                {
                    "module": module,
                    "factor": rec.get("anchor_factor"),
                    "factor_label": rec.get("anchor_factor"),
                    "role": role,
                    "trait_code": rec.get("trait"),
                    "biomarker_name": rec.get("biomarker_name"),
                    "group": rec.get("group"),
                    "used_in_factor_model": "yes" if role == "final8_indicator" else "no",
                    "used_only_for_internal_validation": "no" if role == "final8_indicator" else "yes",
                }
            )
    factor_order = {
        "F1_TG_rich_axis": 1,
        "F2_HDL_core_axis": 2,
        "F3_CE_structural_axis": 3,
        "F1_ketone_axis": 4,
        "F2_amino_acid_axis": 5,
        "F3_energy_bridge_axis": 6,
    }
    role_order = {"final8_indicator": 0, "same_domain_support": 1}
    records.sort(key=lambda r: (r["module"], factor_order.get(r["factor"], 99), role_order.get(r["role"], 9), r["trait_code"]))
    write_records(out_wb, "Guide_FactorTrait_map", header, records)


def write_mixer_focal(out_wb):
    ldsc_header, ldsc_rows = read_tsv_records(NDD_LDSC)
    ldsc_by_pair = {(r["trait1"], r["trait2"]): r for r in ldsc_rows}
    records = []
    for path in MIXER_SUMMARIES:
        _, rows = read_tsv_records(path)
        for m in rows:
            key = (m["trait1"], m["trait2"])
            l = ldsc_by_pair.get(key, {})
            rec = {
                "trait1": m.get("trait1"),
                "trait2": m.get("trait2"),
                "ldsc_rg": l.get("rg"),
                "ldsc_rg_se": l.get("rg_se"),
                "ldsc_p_rg": l.get("p_rg"),
                "ldsc_fdr_rg": l.get("fdr_rg"),
                "ldsc_intercept": l.get("intercept"),
                "mixer_rg": m.get("rg"),
                "mixer_rho_beta": m.get("rho_beta"),
                "pi1": m.get("pi1"),
                "pi2": m.get("pi2"),
                "pi12": m.get("pi12"),
                "pi1u": m.get("pi1u"),
                "pi2u": m.get("pi2u"),
                "dice": m.get("dice"),
                "nc1": m.get("nc1"),
                "nc2": m.get("nc2"),
                "nc12": m.get("nc12"),
                "pi12_over_pi1u": m.get("pi12_over_pi1u"),
                "pi12_over_pi2u": m.get("pi12_over_pi2u"),
                "pi12_over_totalpi": m.get("pi12_over_totalpi"),
                "source": m.get("source"),
            }
            records.append(rec)
    header = [
        "trait1",
        "trait2",
        "ldsc_rg",
        "ldsc_rg_se",
        "ldsc_p_rg",
        "ldsc_fdr_rg",
        "ldsc_intercept",
        "mixer_rg",
        "mixer_rho_beta",
        "pi1",
        "pi2",
        "pi12",
        "pi1u",
        "pi2u",
        "dice",
        "nc1",
        "nc2",
        "nc12",
        "pi12_over_pi1u",
        "pi12_over_pi2u",
        "pi12_over_totalpi",
        "source",
    ]
    write_records(out_wb, "S24_MiXeR_focal_clean", header, records)


def build_contents(out_wb, entries):
    ws = out_wb.create_sheet("Contents", 0)
    ws.append(["table", "sheet_name", "description", "source"])
    for number, sheet_name, description, source in entries:
        ws.append([f"Supplementary Table S{number}", sheet_name, description, source])
    for cell in ws[1]:
        cell.fill = TITLE_FILL
        cell.font = TITLE_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx in range(1, 5):
        ws.column_dimensions[get_column_letter(col_idx)].width = [24, 30, 80, 42][col_idx - 1]


def main():
    comb_wb = load_workbook(COMBINED, data_only=True)
    v5_wb = load_workbook(V5, data_only=True)
    out_wb = Workbook()
    out_wb.remove(out_wb.active)

    entries = []

    def add_from_comb(number, title, src, desc):
        clone_sheet(comb_wb[src], out_wb, title)
        entries.append((number, title, desc, f"combined workbook: {src}"))

    def add_from_v5(number, title, src, desc):
        clone_sheet(v5_wb[src], out_wb, title)
        entries.append((number, title, desc, f"v5 downstream workbook: {src}"))

    add_from_comb(1, "S01_AllMetab_QC_LDSC", "02_Main112_QC_Shared", "All 112 non-proportion metabolic traits with quality-control and univariate LDSC metrics.")
    merge_final16(comb_wb, out_wb)
    entries.append((2, "S02_Final16_details", "Final lipid8 and nonlipid8 trait manifest merged with trait-level LDSC metrics.", "14_FinalManifest_Combined + 15_FinalLDSC_Combined"))
    add_from_comb(3, "S03_Trait_lineage", "05_Lineage_Combined", "Trait lineage from the full metabolic universe to lipid/nonlipid candidate and final models.")
    add_from_comb(4, "S04_StageCriteria", "01_StageCriteria", "Prespecified stage criteria used during metabolic trait curation and factor-model refinement.")
    write_factor_trait_map(out_wb)
    write_matrix_from_long(comb_wb, out_wb, "16_FinalS_Long", "s_estimate", "S05_FinalTrait_S_matrix")
    entries.append((5, "S05_FinalTrait_S_matrix", "Matrix-form multivariable LDSC genetic covariance estimates among retained final-model traits.", "16_FinalS_Long"))
    write_matrix_from_long(comb_wb, out_wb, "17_Finalrg_Long", "rg", "S06_FinalTrait_rg_matrix")
    entries.append((6, "S06_FinalTrait_rg_matrix", "Matrix-form pairwise genetic correlations among retained final-model traits.", "17_Finalrg_Long"))
    add_from_comb(7, "S07_FormalSEM_fit", "31_FormalGenomicSEM_Modelfit", "Formal GenomicSEM model-fit indices for the final lipid8 and nonlipid8 models.")
    add_from_comb(8, "S08_Loadings_all6F", "20_FinalLoadings_Combined", "Factor loadings for all three lipid factors and all three nonlipid factors.")
    merge_factor_gwas_neff(comb_wb, out_wb, "S09_FactorGWAS_Neff")
    entries.append((9, "S09_FactorGWAS_Neff", "Factor GWAS file manifest and effective sample-size summaries for all six factors.", "21_FactorGWAS_Files + 30_FactorGWAS_Neff"))
    add_from_comb(10, "S10_Factor_LDSC_uni", "22_FactorLDSC_Uni", "Univariate LDSC QC estimates for all six factor GWAS, including mean chi-square, lambda GC, LDSC intercept, ratio, and SNP heritability.")
    add_from_comb(11, "S11_Factor_LDSC_bi", "23_FactorLDSC_Bi", "Internal bivariate LDSC between factors within the lipid and nonlipid models.")
    write_internal_bivar_compact(out_wb)
    entries.append((12, "S12_IntFactorMetab_compact", "Factor-centered internal validation table showing LDSC results between each factor GWAS and its primary or same-domain support metabolic traits.", "step24/25_internal_bivariate_*_final8 requested-pairs TSVs"))
    write_internal_bivar(out_wb, "S13_IntFactorMetab_summary", "S14_IntFactorMetab_pairs", "S15_IntValidation_manifest")
    entries.append((13, "S13_IntFactorMetab_summary", "Summary of internal bivariate LDSC between each factor GWAS and its primary, non-primary, and same-domain metabolic traits.", "step24/25_internal_bivariate_*_final8 pairset summary TSVs"))
    entries.append((14, "S14_IntFactorMetab_pairs", "All requested internal bivariate LDSC pairs between factor GWAS and constituent/support metabolic traits.", "step24/25_internal_bivariate_*_final8 requested-pairs TSVs"))
    entries.append((15, "S15_IntValidation_manifest", "Manifest of internal metabolic traits used for factor-metabolite LDSC validation.", "step24/25_internal_bivariate_*_final8 trait manifest TSVs"))
    add_from_comb(16, "S16_QSNP_summary_all6", "33_QSNP_LDClump_Summary", "Genome-wide factor and Q_SNP signal summary for all six factor GWAS.")
    add_from_comb(17, "S17_FactorLeadLoci", "34_FactorLeadLoci_LDClump", "LD-clumped lead loci for all six factor GWAS.")
    add_from_comb(18, "S18_QSNP_LeadLoci", "35_QSNP_LeadLoci_LDClump", "LD-clumped Q_SNP lead loci for all six factor GWAS.")
    add_from_comb(19, "S19_Factor_QSNP_overlap", "36_Factor_QSNP_LD_Overlap", "Overlap between factor lead loci and Q_SNP lead loci.")
    write_tsv_sheet(out_wb, "S20_LOO_sensitivity", LOO_TSV)
    entries.append((20, "S20_LOO_sensitivity", "Leave-one-out sensitivity checks for the final lipid8 model.", "loo_sensitivity/final8_loo_sensitivity_summary.tsv"))
    add_from_comb(21, "S21_ExtData_QC_LDSC", "27_ExtData_QC_UniLDSC", "External metabolite GWAS quality-control and univariate LDSC metrics.")
    add_from_comb(22, "S22_ExtPairSelection", "28_ExtPairSelection", "External metabolite pair selection used for factor-label validation.")
    add_from_comb(23, "S23_ExtBivar_LDSC", "29_ExtBivariateLDSC", "External bivariate LDSC validation between six factor GWAS and selected external metabolite GWAS.")
    write_ndd_ldsc(out_wb, "S24_NDD_LDSC_all18")
    entries.append((24, "S24_NDD_LDSC_all18", "All six metabolic factor GWAS tested against AD, PD, and LBD, including LDSC intercepts and FDR-negative comparisons.", "03_ldsc_metabolic_factors_vs_ndd/metabolic_factors_vs_ndd_requested_pairs.tsv"))
    write_mixer_focal(out_wb)
    out_wb["S24_MiXeR_focal_clean"].title = "S25_MiXeR_focal_clean"
    entries.append((25, "S25_MiXeR_focal_clean", "MiXeR polygenic overlap estimates for the three FDR-selected factor-disease pairs, merged with LDSC rg and intercept.", "05_mixer_*_summary.tsv + NDD LDSC TSV"))

    downstream = [
        (26, "S26_PleioFDR_summary", "S14_PleioFDR_summary", "conjFDR/condFDR summary for focal factor-disease pairs."),
        (27, "S27_PleioFDR_loci", "S15_PleioFDR_loci", "Pleiotropic loci from conjFDR/condFDR analyses."),
        (28, "S28_Coloc_summary", "S16_Coloc_summary", "Colocalization summary for pleiotropic loci."),
        (29, "S29_PWCoCo_best", "S17_PWCoCo_best", "PWCoCo evidence for prioritized pleiotropic regions."),
        (30, "S30_SNP_evidence", "S19_SNP_evidence", "SNP-level evidence integrating pleiotropy, colocalization, and annotation layers."),
        (31, "S31_FUMA_mapping", "S18_FUMA_mapping_summary", "FUMA mapped genes and positional/eQTL/chromatin mapping evidence."),
        (32, "S32_cTWAS_overview", "S20_cTWAS_overview", "cTWAS overview for focal factor-disease pairs."),
        (33, "S33_Candidate_tiers", "S22_Candidate_tier_summary", "Candidate-gene tier summary."),
        (34, "S34_Candidate_master", "S21_Candidate_master", "Master candidate-gene table across focal factor-disease pairs."),
        (35, "S35_Core_genes", "S21a_Candidate_core_genes", "Core prioritized genes used for downstream interpretation and target checks."),
        (36, "S36_BulkSMR_summary", "S23_BulkBrain_SMR_summary", "Bulk brain SMR summary."),
        (37, "S37_BulkSMR_core", "S23a_BulkSMR_core_gene_level", "Core-gene bulk brain SMR evidence."),
        (38, "S38_GTExSMR_summary", "S24_GTEx_SMR_summary", "GTEx SMR summary."),
        (39, "S39_GTExSMR_core", "S24a_GTExSMR_core_gene_level", "Core-gene GTEx SMR evidence."),
        (40, "S40_CellTypeSMR_summary", "S25_CellType_SMR_summary", "Cell-type SMR summary."),
        (41, "S41_CellTypeSMR_core", "S25a_BryoisSMR_core_gene_level", "Core-gene cell-type SMR evidence."),
        (42, "S42_scPagwas_overview", "S27_scPagwas_overview", "scPagwas overview for focal factor-disease pairs."),
        (43, "S43_scPagwas_celltypes", "S28_F2_AD_celltypes", "scPagwas cell-type enrichment results."),
        (44, "S44_scPagwas_pathways", "S29_F2_AD_pathways", "scPagwas pathway enrichment results."),
        (45, "S45_scPagwas_curated", "S29a_scPagwas_pathway_curated", "Curated scPagwas pathway evidence."),
        (46, "S46_PCC_curated", "S30_PCC_curated", "Curated protein/coexpression context evidence."),
        (47, "S47_KNK_summary_all", "S31a_KNK_summary_all", "KNK-prioritized gene summary across focal analyses."),
        (48, "S48_KNK_overlap_all", "S31b_KNK_overlap_all", "KNK overlap evidence across prioritized genes and pathways."),
        (49, "S49_KNK_pathway_summary", "S33_KNK_pathway_summary", "KNK pathway summary."),
        (50, "S50_KNK_target_genes", "S39_KNK_target_genes", "KNK-entering target genes taken forward for target annotation."),
        (51, "S51_Target_tractability", "S40_Target_tractability", "Open Targets tractability and modality annotations."),
        (52, "S52_Brain_disease_support", "S42_Brain_disease_support", "Human Protein Atlas/Open Targets brain and disease support annotations."),
        (53, "S53_Binding_modality", "S43_Binding_modality", "Small-molecule/protein-binding modality annotations for target genes."),
        (54, "S54_Target_PPI_detail", "S47_Target_PPI_detail", "Protein-protein interaction detail for target genes."),
        (55, "S55_Chemical_probes", "S48_Chemical_probes", "Chemical-probe and ligand evidence for prioritized targets."),
        (56, "S56_RouteA_external", "S44_RouteA_external", "Public disease-state dataset manifest for supportive external checks."),
        (57, "S57_PD_meta_signature", "S45_PD_meta_signature", "Supportive PD transcriptomic meta-signature checks for KNK-prioritized genes."),
        (58, "S58_AD_CSF_proteomics", "S46_AD_CSF_proteomics", "Supportive AD CSF proteomics check for available KNK-prioritized genes."),
        (59, "S59_Audit_triplet", "Audit_triplet", "Audit trail for focal triplet supplementary analyses."),
        (60, "S60_Audit_submission", "Audit_submission_slim", "Submission-slim workbook audit trail."),
        (61, "S61_Audit_target_validation", "Audit_target_validation", "Target-validation audit trail."),
    ]
    for number, title, src, desc in downstream:
        add_from_v5(number, title, src, desc)

    # Keep process-heavy tables at the end as audit/method-support sheets, not as
    # primary result tables.
    process_sheets = [
        (62, "S62_scPagwas_atlas", "S26_scPagwas_atlas", "scPagwas atlas metadata retained as methodological provenance."),
        (63, "S63_KNK_overview", "S31_KNK_overview", "KNK run/provenance overview retained as methodological provenance."),
        (64, "S64_KNK_core4_summary", "S32_KNK_core4_summary", "Legacy KNK core4 summary retained for traceability."),
        (65, "S65_KNK_pericyte_gene", "S34_KNK_pericyte_gene_overlap", "Legacy pericyte gene-overlap output retained for traceability."),
        (66, "S66_KNK_pericyte_path", "S35_KNK_pericyte_path_overlap", "Legacy pericyte pathway-overlap output retained for traceability."),
        (67, "S67_KNK_opc_gene", "S36_KNK_opc_gene_overlap", "Legacy OPC gene-overlap output retained for traceability."),
        (68, "S68_KNK_opc_path", "S37_KNK_opc_path_overlap", "Legacy OPC pathway-overlap output retained for traceability."),
        (69, "S69_KNK_scPagwas_overlap", "S38_KNK_scPagwas_overlap", "Legacy KNK-scPagwas overlap output retained for traceability."),
        (70, "S70_Factor_annotation", "S41_Factor_annotation", "Focal-factor annotation notes retained for traceability."),
    ]
    for number, title, src, desc in process_sheets:
        add_from_v5(number, title, src, desc)

    build_contents(out_wb, entries)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(OUT)
    print(OUT)


if __name__ == "__main__":
    main()

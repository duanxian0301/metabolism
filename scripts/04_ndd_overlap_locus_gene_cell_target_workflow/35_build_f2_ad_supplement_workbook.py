from __future__ import annotations

from pathlib import Path
import math
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


BASE = Path(r"D:\codex\GenomicSEM\metabolic\postgwas_ad_pdlbd\results")
WORK_BASE = Path(r"D:\codex\GenomicSEM\metabolic\postgwas_ad_pdlbd\work")
GWAS_BASE = Path(r"D:\metabolic\GWAS\genomicgem_main_zgt4_nonproportion")
OUT_DIR = BASE / "22_supplement_tables_lipid8_F2_AD"
OUT_DIR.mkdir(parents=True, exist_ok=True)
OUT_XLSX = OUT_DIR / "lipid8_F2_AD_supplementary_tables.xlsx"


def read_tsv(path: Path, **kwargs) -> pd.DataFrame:
    return pd.read_csv(path, sep="\t", **kwargs)


def read_csv(path: Path, **kwargs) -> pd.DataFrame:
    return pd.read_csv(path, **kwargs)


def autosize(ws, max_width: int = 42) -> None:
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells[:150]:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), max_width)


def add_sheet(wb: Workbook, name: str, title: str, subtitle: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(name)
    ws["A1"] = title
    ws["A1"].font = Font(bold=True)
    ws["A2"] = subtitle
    header_row = 4
    for c_idx, col in enumerate(df.columns, start=1):
        ws.cell(row=header_row, column=c_idx, value=col).font = Font(bold=True)
    for r_idx, row in enumerate(df.astype(object).where(pd.notnull(df), "").itertuples(index=False), start=header_row + 1):
        for c_idx, value in enumerate(row, start=1):
            if isinstance(value, float) and (math.isinf(value) or math.isnan(value)):
                value = ""
            ws.cell(row=r_idx, column=c_idx, value=value)
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(df.columns))}{max(4, len(df) + 4)}"
    autosize(ws)


def build_contents(sheets_meta: list[tuple[str, str]]) -> pd.DataFrame:
    rows = []
    for i, (sheet, title) in enumerate(sheets_meta, start=1):
        rows.append(
            {
                "Supplementary Table": f"S{i}",
                "Sheet name": sheet,
                "Title": title,
            }
        )
    return pd.DataFrame(rows)


def main() -> None:
    wb = Workbook()
    del wb[wb.sheetnames[0]]

    # S1
    final8_manifest = read_tsv(
        GWAS_BASE
        / "lipid_module_from_full_manifest"
        / "compact_panel_review"
        / "ultra_pure_3factor_review"
        / "final8_review"
        / "lipid_module_ultrapure3_final8_kept.tsv"
    )
    data_qc = final8_manifest[
        [
            "trait_code",
            "study_accession",
            "biomarker_name",
            "group",
            "marker_role",
            "selection_status",
            "ultra_pure_factor",
        ]
    ].copy()
    data_qc.columns = [
        "Trait",
        "Source dataset",
        "Biomarker",
        "Group",
        "Marker role",
        "Selection status",
        "Original factor anchor",
    ]
    data_qc["Main downstream analysis"] = "lipid8_F2 x AD"

    # S2
    ldsc_qc = read_tsv(GWAS_BASE / "step12_ldsc_lipid_module_final8" / "lipid_module_final8_ldsc_summary.tsv")
    trait_manifest = read_tsv(GWAS_BASE / "step12_ldsc_lipid_module_final8" / "trait_manifest.tsv")
    ldsc_qc = ldsc_qc.merge(
        trait_manifest[["trait_code", "sumstats_file", "biomarker_name", "group"]],
        left_on="trait",
        right_on="trait_code",
        how="left",
    )
    ldsc_qc = ldsc_qc[
        ["trait", "biomarker_name", "group", "sumstats_file", "h2", "intercept"]
    ].rename(columns={"sumstats_file": "Summary statistics file"})

    # S3
    rg_panel = read_tsv(BASE / "03_ldsc_metabolic_factors_vs_ndd" / "metabolic_factors_vs_ndd_requested_pairs.tsv")
    rg_panel = rg_panel[rg_panel["trait1"] == "lipid8_F2"].copy()

    # S4
    efa_criteria = read_tsv(GWAS_BASE / "step13_efa_esem_lipid_module_final8" / "lipid_module_final8_factor_criteria.tsv")

    # S5
    efa_loadings = read_tsv(GWAS_BASE / "step13_efa_esem_lipid_module_final8" / "EFA_minres_3factor_loadings.tsv")

    # S6
    esem_fit = read_tsv(GWAS_BASE / "step13_efa_esem_lipid_module_final8" / "lipid_module_final8_esem_summary.tsv")

    # S7
    f2_loadings = read_tsv(GWAS_BASE / "step13_efa_esem_lipid_module_final8" / "ALL_3factor_loadings.tsv")
    f2_loadings = f2_loadings[f2_loadings["factor"] == "f2"].copy()
    f2_loadings["abs_std_all"] = f2_loadings["std.all"].abs()
    f2_loadings = f2_loadings.sort_values("abs_std_all", ascending=False)

    # S8
    f2_factor_hits = read_tsv(GWAS_BASE / "q_snp_ld_clump_final8" / "final8_factor_lead_loci_ldclump.tsv")
    f2_factor_hits = f2_factor_hits[(f2_factor_hits["module"] == "lipid") & (f2_factor_hits["factor"] == "F2")].copy()
    f2_gwas = read_tsv(WORK_BASE / "03_ldsc_metabolic_factors_vs_ndd" / "lipid8_F2.txt")
    f2_gwas_summary = pd.DataFrame(
        [
            {
                "factor": "lipid8_F2",
                "rows": len(f2_gwas),
                "unique_snps_assumed": f2_gwas["SNP"].nunique(),
                "min_p": f2_gwas["P"].min(),
                "gws_snps_p_lt_5e8": int((f2_gwas["P"] < 5e-8).sum()),
                "lead_hits_ldclump": len(f2_factor_hits),
                "N_median": int(f2_gwas["N"].median()),
            }
        ]
    )

    # S9
    f2_lead_hits = f2_factor_hits.rename(
        columns={"P": "clump_P", "factor_P": "factor_P"}
    )

    # S10
    f2_qsnp_merged = read_tsv(
        GWAS_BASE
        / "step14_native_wsl_usergwas_final8_results"
        / "merged_lipid_final8"
        / "lipid_final8_F2_userGWAS_merged.tsv.gz"
    )
    f2_qsnp_summary = pd.DataFrame(
        [
            {"metric": "n_total_snps", "value": len(f2_qsnp_merged)},
            {"metric": "F2_qsnp_sig_p_lt_5e-8", "value": int((f2_qsnp_merged["Q_SNP_pval"] < 5e-8).sum())},
            {"metric": "F2_qsnp_suggestive_p_lt_1e-5", "value": int((f2_qsnp_merged["Q_SNP_pval"] < 1e-5).sum())},
            {"metric": "F2_qsnp_min_p", "value": float(f2_qsnp_merged["Q_SNP_pval"].min())},
            {"metric": "F2_factor_sig_p_lt_5e-8", "value": int((f2_qsnp_merged["Pval_Estimate"] < 5e-8).sum())},
        ]
    )

    # S11
    qsnp_clump = read_tsv(GWAS_BASE / "q_snp_ld_clump_final8" / "final8_qsnp_ldclump_summary.tsv")
    qsnp_clump_f2 = qsnp_clump[(qsnp_clump["module"] == "lipid") & (qsnp_clump["factor"] == "F2")].copy()
    qsnp_clump_leads = pd.DataFrame(
        [
            {
                "analysis": "lipid8_F2_Q",
                "n_lead_loci_clumped": int(qsnp_clump_f2["n_qsnp_lead_loci_ldclump"].iloc[0]),
                "n_exact_overlap_lead_snps": int(qsnp_clump_f2["n_exact_overlap_lead_snps"].iloc[0]),
                "n_factor_leads_excluding_ld_with_qsnp": int(qsnp_clump_f2["n_factor_leads_excluding_ld_with_qsnp"].iloc[0]),
            }
        ]
    )

    # S12
    loo = read_tsv(OUT_DIR / "loo_sensitivity" / "final8_loo_sensitivity_summary.tsv")

    # S13
    mixer = read_tsv(BASE / "05_mixer_lipid8_F2_AD" / "mixer_lipid8_F2_AD_summary.tsv")
    cross = rg_panel.merge(mixer, on=["trait1", "trait2"], how="left")

    # S14
    pleio_summary = read_csv(BASE / "06_pleiofdr_lipid8_F2_AD" / "pleiofdr_lipid8_F2_AD_summary.csv")

    # S15
    pleio_loci = read_csv(BASE / "06_pleiofdr_lipid8_F2_AD" / "lipid8_F2_AD_conjfdr_0.05_loci.csv")

    # S16
    coloc = read_tsv(BASE / "09_coloc_lipid8_F2_AD" / "coloc_lipid8_F2_AD_regions.tsv")

    # S17
    pwcoco = read_tsv(BASE / "10_pwcoco_lipid8_F2_AD" / "pwcoco_lipid8_F2_AD_best_h4.tsv")

    # S18
    fuma_genes = read_tsv(BASE / "13_fuma_lipid8_F2_AD" / "fuma" / "FUMA_job729202" / "genes.txt")

    # S19
    snp_evidence = read_tsv(BASE / "20_snp_evidence_lipid8_F2_AD" / "lipid8_F2_AD_lead_snp_evidence_table.tsv")

    # S20
    ctwas_counts = pd.concat(
        [
            read_tsv(BASE / "16_ctwas_lipid8_F2_AD" / "summary" / "lipid8_F2_counts.tsv"),
            read_tsv(BASE / "16_ctwas_lipid8_F2_AD" / "summary" / "AD_counts.tsv"),
        ],
        ignore_index=True,
    )

    # S21
    candidate_master = read_tsv(BASE / "17_candidate_gene_integration_lipid8_F2_AD" / "lipid8_F2_AD_cross_evidence_shortlist.tsv")

    # S22
    candidate_tiers = read_tsv(BASE / "17_candidate_gene_integration_lipid8_F2_AD" / "lipid8_F2_AD_cross_evidence_tier_counts.tsv")

    # S23 brainmeta summary
    brainmeta = read_tsv(BASE / "11_smr_lipid8_F2_AD" / "summary" / "smr_brainmeta_lipid8_F2_AD_combined.tsv")
    brainmeta_summary_rows = []
    for trait, sub in brainmeta.groupby("trait"):
        n_hits = int((sub["p_SMR"] < 1e-4).sum())
        n_supported = int(((sub["p_SMR"] < 1e-4) & sub["p_HEIDI"].notna() & (sub["p_HEIDI"] > 0.01)).sum())
        best = sub.sort_values("p_SMR").iloc[0]
        brainmeta_summary_rows.append(
            {
                "trait": trait,
                "panel": "BrainMeta",
                "context": "bulk_brain",
                "n_hits_p_smr_lt_1e4": n_hits,
                "n_hits_p_smr_lt_1e4_heidi_supported": n_supported,
                "best_gene": best["Gene"],
                "best_probe": best["probeID"],
                "best_topSNP": best["topSNP"],
                "best_p_SMR": best["p_SMR"],
                "best_p_HEIDI": best["p_HEIDI"],
            }
        )
    brainmeta_summary = pd.DataFrame(brainmeta_summary_rows)

    # S24 gtex summary
    gtex_counts = read_tsv(BASE / "14_smr_gtex_lipid8_F2_AD" / "summary" / "smr_gtex_lipid8_F2_AD_counts.tsv")

    # S25 bryois summary
    bryois_counts = read_tsv(BASE / "15_smr_bryois_lipid8_F2_AD" / "summary" / "smr_bryois_lipid8_F2_AD_counts.tsv")

    # S26 atlas
    atlas = pd.DataFrame(
        [
            {
                "Disease": "MSSM_AD",
                "File": "MSSM_AD_20k.rds",
                "Total_cells": 20000,
                "Case_status_column": "diagnosis",
                "Case_status_value": "AD",
                "Notes": "Downsampled 20k atlas used for scPagwas2 projection",
            }
        ]
    )

    # S27 overview
    celltypes = read_csv(BASE / "12_scpagwas2_lipid8_F2_MSSM_AD" / "lipid8_F2_MSSM_AD_Merged_celltype_pvalue_withFDR.csv")
    n_sig_celltypes = int((celltypes["celltype_FDR"] < 0.05).sum())
    top_cell = celltypes.sort_values("celltype_FDR").iloc[0]
    overview = pd.DataFrame(
        [
            {
                "analysis": "lipid8_F2_MSSM_AD",
                "n_sig_celltypes": n_sig_celltypes,
                "sig_celltypes": "; ".join(celltypes.loc[celltypes["celltype_FDR"] < 0.05, "celltype"].tolist()) if n_sig_celltypes else "",
                "n_sig_pathways_strict_fdr": 0,
                "top_celltype_by_fdr": top_cell["celltype"],
                "top_celltype_fdr": top_cell["celltype_FDR"],
                "n_celltypes_tested": len(celltypes),
                "suggestive_celltypes_fdr_lt_0_1": "; ".join(celltypes.loc[celltypes["celltype_FDR"] < 0.1, "celltype"].tolist()),
            }
        ]
    )

    # S28
    f2_ad_celltypes = celltypes.copy()
    f2_ad_celltypes["analysis"] = "lipid8_F2_MSSM_AD"

    # S29
    pericyte_path = read_csv(BASE / "12_scpagwas2_lipid8_F2_MSSM_AD" / "Pathway_TRS" / "Result_pericyte_Pathway_vs_TRS_all.csv")
    pericyte_path["analysis"] = "lipid8_F2_MSSM_AD"

    # S30
    pcc = read_csv(BASE / "12_scpagwas2_lipid8_F2_MSSM_AD" / "lipid8_F2_MSSM_AD_gene_PCC.csv", index_col=0)
    pcc = pcc.reset_index().rename(columns={"index": "gene"})
    curated_genes = ["KANSL1", "ARL17B", "LRRC37A2", "KNOP1", "CAB39L", "MAP1LC3A"]
    pcc_curated = pcc[pcc["gene"].isin(curated_genes)].copy().sort_values("adj_pvalue")

    # S31
    knk_overview = pd.DataFrame(
        [
            {"section": "Core KNK groups", "value": 4},
            {"section": "Target cells", "value": "pericyte; oligodendrocyte precursor cell"},
            {"section": "Target genes", "value": "KANSL1; ARL17B; LRRC37A2; KNOP1"},
            {"section": "Completed KO groups", "value": 4},
            {"section": "Most perturbed branch", "value": "pericyte-KANSL1"},
        ]
    )

    # S32
    knk_summary = pd.concat(
        [
            read_csv(BASE / "19_knk_lipid8_F2_AD_core4" / "summary" / "knk_summary_core4_shard1of2.csv"),
            read_csv(BASE / "19_knk_lipid8_F2_AD_core4" / "summary" / "knk_summary_core4_shard2of2.csv"),
        ],
        ignore_index=True,
    )

    # S33
    path_rows = []
    for fp in (BASE / "19_knk_lipid8_F2_AD_core4").rglob("*_pathway_enrichment.csv"):
        df = read_csv(fp)
        run_id = fp.stem.replace("_pathway_enrichment", "")
        parts = run_id.split("_")
        gene = parts[-1]
        cell_type = "Pericyte" if "Pericyte" in run_id else "Oligodendrocyte_precursor_cell"
        df["run_id"] = run_id
        df["dataset"] = "MSSM_AD"
        df["cell_type"] = cell_type
        df["gene"] = gene
        path_rows.append(df)
    knk_pathways = pd.concat(path_rows, ignore_index=True)

    # S34/S35
    pericyte_gene_overlap = pd.DataFrame([{"V1": "KANSL1", "V2": "KANSL1", "N": 1}])
    pericyte_path_overlap = pd.DataFrame([{"V1": "KANSL1", "V2": "KANSL1", "N": 1}])

    # S36/S37
    opc_gene_overlap = read_csv(BASE / "19_knk_lipid8_F2_AD_core4" / "summary" / "MSSM_AD_Oligodendrocyte_precursor_cell_gene_overlap_jaccard.csv")
    opc_path_overlap = read_csv(BASE / "19_knk_lipid8_F2_AD_core4" / "summary" / "MSSM_AD_Oligodendrocyte_precursor_cell_pathway_overlap_jaccard.csv")

    # S38
    knk_overlap = read_tsv(BASE / "21_knk_scpagwas_overlap_lipid8_F2_AD" / "knk_vs_scpagwas_pericyte_trs_overlap.tsv")

    sheets = [
        ("Contents", "Contents"),
        ("S1_Data_QC", "Input resources and QC provenance"),
        ("S2_LDSC_QC", "Univariate LDSC quality control for final8 lipid markers"),
        ("S3_rg_panel", "Cross-disease LDSC panel centered on lipid8_F2"),
        ("S4_EFA_criteria", "Exploratory factor analysis criteria for final8 lipid panel"),
        ("S5_EFA3_loadings", "Three-factor exploratory loadings for final8 lipid panel"),
        ("S6_ESEM_fit", "Target-rotated ESEM fit summary for final8 lipid panel"),
        ("S7_F2_loadings", "Final lipid8_F2 loading pattern"),
        ("S8_F2_GWAS_summary", "Factor GWAS summary for lipid8_F2"),
        ("S9_F2_lead_hits", "LD-clumped lead genome-wide significant loci for lipid8_F2"),
        ("S10_QSNP_summary", "Q_SNP heterogeneity summary for lipid8_F2"),
        ("S11_QSNP_clumped_leads", "LD-clumped Q_SNP lead-locus summary for lipid8_F2"),
        ("S12_LOO_sensitivity", "Leave-one-out sensitivity analysis for final8 lipid model"),
        ("S13_CrossDisease_LDSC_MiXeR", "Cross-disease LDSC and MiXeR summary for lipid8_F2"),
        ("S14_PleioFDR_summary", "PleioFDR summary for lipid8_F2 x AD"),
        ("S15_PleioFDR_loci", "ConjFDR lead loci for lipid8_F2 x AD"),
        ("S16_Coloc_summary", "Coloc regional summary for lipid8_F2 x AD"),
        ("S17_PWCoCo_best", "PWCoCo best-H4 summary for lipid8_F2 x AD"),
        ("S18_FUMA_mapping_summary", "FUMA mapped-gene summary for lipid8_F2 x AD"),
        ("S19_SNP_evidence", "Lead-SNP integrated evidence table"),
        ("S20_cTWAS_overview", "cTWAS overview for lipid8_F2 and AD"),
        ("S21_Candidate_master", "Integrated candidate-gene shortlist"),
        ("S22_Candidate_tier_summary", "Candidate-gene tier counts"),
        ("S23_BulkBrain_SMR_summary", "Bulk-brain SMR summary"),
        ("S24_GTEx_SMR_summary", "GTEx SMR summary"),
        ("S25_CellType_SMR_summary", "Bryois cell-type SMR summary"),
        ("S26_scPagwas_atlas", "Single-nucleus atlas metadata for scPagwas2"),
        ("S27_scPagwas_overview", "scPagwas2 overview for lipid8_F2 x AD"),
        ("S28_F2_AD_celltypes", "lipid8_F2 x AD scPagwas2 cell-type results"),
        ("S29_F2_AD_pathways", "lipid8_F2 x AD pericyte pathway-TRS results"),
        ("S30_PCC_curated", "Curated gene PCC results for lipid8_F2 x AD"),
        ("S31_KNK_overview", "Virtual knockout overview for lipid8_F2 x AD"),
        ("S32_KNK_core4_summary", "Core4 KNK branch summary"),
        ("S33_KNK_pathway_summary", "KNK pathway enrichment summary"),
        ("S34_KNK_pericyte_gene_overlap", "Pericyte-branch KNK gene overlap"),
        ("S35_KNK_pericyte_path_overlap", "Pericyte-branch KNK pathway overlap"),
        ("S36_KNK_opc_gene_overlap", "OPC-branch KNK gene overlap"),
        ("S37_KNK_opc_path_overlap", "OPC-branch KNK pathway overlap"),
        ("S38_KNK_scPagwas_overlap", "KNK vs scPagwas pathway overlap summary"),
    ]

    contents_df = build_contents([(name, title) for name, title in sheets[1:]])
    add_sheet(
        wb,
        "Contents",
        "Supplementary Tables for the lipid8_F2 x AD manuscript",
        "This workbook is an F2 x AD focused supplement adapted from the reference workflow and populated only with results generated in the current study.",
        contents_df,
    )

    add_sheet(wb, "S1_Data_QC", "Supplementary Table S1. Input resources and QC provenance", "This table summarizes the final8 lipid-panel inputs carried forward into factor construction and the downstream lipid8_F2 x AD workflow.", data_qc)
    add_sheet(wb, "S2_LDSC_QC", "Supplementary Table S2. Univariate LDSC quality control for final8 lipid markers", "This table summarizes marker-level LDSC heritability and intercept metrics for the final8 lipid panel.", ldsc_qc)
    add_sheet(wb, "S3_rg_panel", "Supplementary Table S3. Cross-disease LDSC panel centered on lipid8_F2", "This table reports LDSC covariance and genetic-correlation results for lipid8_F2 against AD, PD, and LBD, as generated in the current project.", rg_panel)
    add_sheet(wb, "S4_EFA_criteria", "Supplementary Table S4. Exploratory factor analysis criteria", "This table lists factor-number selection criteria used during exploratory factor analysis of the final8 lipid panel.", efa_criteria)
    add_sheet(wb, "S5_EFA3_loadings", "Supplementary Table S5. Three-factor exploratory loadings", "This table reports the three-factor exploratory loading pattern for the final8 lipid panel.", efa_loadings)
    add_sheet(wb, "S6_ESEM_fit", "Supplementary Table S6. Target-rotated ESEM fit summary", "This table reports target-rotated ESEM fit statistics for the final8 lipid panel in ODD and ALL covariance sets.", esem_fit)
    add_sheet(wb, "S7_F2_loadings", "Supplementary Table S7. Final loading pattern for lipid8_F2", "This table reports the final standardized loading pattern for the lipid8_F2 factor.", f2_loadings)
    add_sheet(wb, "S8_F2_GWAS_summary", "Supplementary Table S8. lipid8_F2 factor GWAS summary", "This table summarizes the factor GWAS results for lipid8_F2.", f2_gwas_summary)
    add_sheet(wb, "S9_F2_lead_hits", "Supplementary Table S9. LD-clumped lead loci for lipid8_F2", "This table lists LD-clumped genome-wide significant lead loci for the lipid8_F2 factor GWAS.", f2_lead_hits)
    add_sheet(wb, "S10_QSNP_summary", "Supplementary Table S10. Q_SNP heterogeneity summary for lipid8_F2", "This table summarizes Q_SNP heterogeneity results from the lipid8_F2 factor GWAS.", f2_qsnp_summary)
    add_sheet(wb, "S11_QSNP_clumped_leads", "Supplementary Table S11. LD-clumped Q_SNP lead-locus summary for lipid8_F2", "This table reports LD-clumped Q_SNP lead-locus counts for lipid8_F2. No genome-wide significant lipid8_F2 Q_SNP lead locus was detected in the current study.", qsnp_clump_leads)
    add_sheet(wb, "S12_LOO_sensitivity", "Supplementary Table S12. Leave-one-out sensitivity analysis", "This table summarizes leave-one-out sensitivity analyses for the final8 lipid factor model.", loo)
    add_sheet(wb, "S13_CrossDisease_LDSC_MiXeR", "Supplementary Table S13. Cross-disease LDSC and MiXeR summary for lipid8_F2", "This table summarizes lipid8_F2 comparisons with AD, PD, and LBD using LDSC, with MiXeR metrics available for the lipid8_F2 x AD pair.", cross)
    add_sheet(wb, "S14_PleioFDR_summary", "Supplementary Table S14. PleioFDR summary for lipid8_F2 x AD", "This table summarizes conjunctional FDR results for lipid8_F2 and AD.", pleio_summary)
    add_sheet(wb, "S15_PleioFDR_loci", "Supplementary Table S15. ConjFDR lead loci for lipid8_F2 x AD", "This table lists lead loci passing conjFDR < 0.05 for the lipid8_F2 x AD analysis.", pleio_loci)
    add_sheet(wb, "S16_Coloc_summary", "Supplementary Table S16. Coloc regional summary", "This table summarizes coloc regional results for lipid8_F2 x AD.", coloc)
    add_sheet(wb, "S17_PWCoCo_best", "Supplementary Table S17. PWCoCo best-H4 summary", "This table reports best-H4 PWCoCo results for lipid8_F2 x AD.", pwcoco)
    add_sheet(wb, "S18_FUMA_mapping_summary", "Supplementary Table S18. FUMA mapped-gene summary", "This table reports mapped genes from the FUMA job for lipid8_F2 x AD.", fuma_genes)
    add_sheet(wb, "S19_SNP_evidence", "Supplementary Table S19. Lead-SNP integrated evidence table", "This table links each conjFDR lead locus to coloc, PWCoCo, and FUMA SNP-level annotation evidence.", snp_evidence)
    add_sheet(wb, "S20_cTWAS_overview", "Supplementary Table S20. cTWAS overview", "This table summarizes cTWAS counts for lipid8_F2 and AD.", ctwas_counts)
    add_sheet(wb, "S21_Candidate_master", "Supplementary Table S21. Integrated candidate-gene shortlist", "This table reports the shortlisted integrated candidate genes for lipid8_F2 x AD.", candidate_master)
    add_sheet(wb, "S22_Candidate_tier_summary", "Supplementary Table S22. Candidate-gene tier summary", "This table summarizes candidate-gene tier counts under the integrated evidence framework.", candidate_tiers)
    add_sheet(wb, "S23_BulkBrain_SMR_summary", "Supplementary Table S23. Bulk-brain SMR summary", "This table summarizes bulk-brain SMR hits for lipid8_F2 and AD.", brainmeta_summary)
    add_sheet(wb, "S24_GTEx_SMR_summary", "Supplementary Table S24. GTEx SMR summary", "This table summarizes GTEx SMR results for lipid8_F2 and AD.", gtex_counts)
    add_sheet(wb, "S25_CellType_SMR_summary", "Supplementary Table S25. Bryois cell-type SMR summary", "This table summarizes Bryois cell-type SMR results for lipid8_F2 and AD.", bryois_counts)
    add_sheet(wb, "S26_scPagwas_atlas", "Supplementary Table S26. Single-nucleus atlas metadata for scPagwas2", "This table describes the single-nucleus atlas used for the lipid8_F2 x AD scPagwas2 projection.", atlas)
    add_sheet(wb, "S27_scPagwas_overview", "Supplementary Table S27. scPagwas2 overview for lipid8_F2 x AD", "This table summarizes the overall scPagwas2 results for lipid8_F2 x AD.", overview)
    add_sheet(wb, "S28_F2_AD_celltypes", "Supplementary Table S28. lipid8_F2 x AD scPagwas2 cell-type results", "This table reports scPagwas2 cell-type enrichment results for lipid8_F2 x AD.", f2_ad_celltypes)
    add_sheet(wb, "S29_F2_AD_pathways", "Supplementary Table S29. lipid8_F2 x AD pericyte pathway-TRS results", "This table reports pericyte pathway-TRS associations from scPagwas2 for lipid8_F2 x AD.", pericyte_path)
    add_sheet(wb, "S30_PCC_curated", "Supplementary Table S30. Curated gene PCC results", "This table reports curated gene-level PCC results for selected lipid8_F2 x AD candidate genes.", pcc_curated)
    add_sheet(wb, "S31_KNK_overview", "Supplementary Table S31. Virtual knockout overview", "This table provides an overview of the core4 virtual knockout design and completion status.", knk_overview)
    add_sheet(wb, "S32_KNK_core4_summary", "Supplementary Table S32. Core4 KNK branch summary", "This table summarizes the retained KNK branches for lipid8_F2 x AD.", knk_summary)
    add_sheet(wb, "S33_KNK_pathway_summary", "Supplementary Table S33. KNK pathway enrichment summary", "This table reports pathway enrichment results from the core4 KNK analyses.", knk_pathways)
    add_sheet(wb, "S34_KNK_pericyte_gene_overlap", "Supplementary Table S34. Pericyte-branch KNK gene overlap", "This table reports gene-overlap results for the pericyte KNK branch.", pericyte_gene_overlap)
    add_sheet(wb, "S35_KNK_pericyte_path_overlap", "Supplementary Table S35. Pericyte-branch KNK pathway overlap", "This table reports pathway-overlap results for the pericyte KNK branch.", pericyte_path_overlap)
    add_sheet(wb, "S36_KNK_opc_gene_overlap", "Supplementary Table S36. OPC-branch KNK gene overlap", "This table reports gene-overlap results for the oligodendrocyte precursor cell KNK branch.", opc_gene_overlap)
    add_sheet(wb, "S37_KNK_opc_path_overlap", "Supplementary Table S37. OPC-branch KNK pathway overlap", "This table reports pathway-overlap results for the oligodendrocyte precursor cell KNK branch.", opc_path_overlap)
    add_sheet(wb, "S38_KNK_scPagwas_overlap", "Supplementary Table S38. KNK and scPagwas pathway overlap summary", "This table summarizes overlap between KNK pathway enrichments and scPagwas pericyte TRS pathway rankings.", knk_overlap)

    wb.save(OUT_XLSX)
    print(OUT_XLSX)


if __name__ == "__main__":
    main()

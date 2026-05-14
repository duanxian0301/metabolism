from __future__ import annotations

import math
import shutil
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


BASE = Path(r"D:\codex\GenomicSEM\metabolic\postgwas_ad_pdlbd\results")
SCP_BASE = Path(r"D:\scPagwas\metabolic_scpagwas2")
V2 = BASE / "22_supplement_tables_lipid8_F2_AD" / "metabolic_factor_triplet_supplementary_tables_v2.xlsx"
OUT = BASE / "22_supplement_tables_lipid8_F2_AD" / "metabolic_factor_triplet_supplementary_tables_v4_slim_submission.xlsx"


PAIR_SPECS = [
    {
        "pair": "lipid8_F2_AD",
        "factor": "lipid8_F2",
        "disease": "AD",
        "label": "lipid8_F2 脳 AD",
        "scp_prefix": "lipid8_F2_MSSM_AD",
        "scp_dir": SCP_BASE / "lipid8_F2_MSSM_AD",
        "knk_summary_paths": [
            BASE / "19_knk_lipid8_F2_AD_core4" / "summary" / "knk_summary_core4_shard1of2.csv",
            BASE / "19_knk_lipid8_F2_AD_core4" / "summary" / "knk_summary_core4_shard2of2.csv",
        ],
        "knk_overlap_paths": [
            BASE / "21_knk_scpagwas_overlap_lipid8_F2_AD" / "knk_vs_scpagwas_pericyte_trs_overlap.tsv",
        ],
    },
    {
        "pair": "nonlipid8_F1_PD",
        "factor": "nonlipid8_F1",
        "disease": "PD",
        "label": "nonlipid8_F1 脳 PD",
        "scp_prefix": "nonlipid8_F1_MSSM_PD",
        "scp_dir": SCP_BASE / "nonlipid8_F1_MSSM_PD",
        "knk_summary_paths": [
            BASE / "24_knk_nonlipid8_F1_PD" / "summary" / "knk_summary_nonlipid8_F1_PD.csv",
        ],
        "knk_overlap_paths": [
            BASE / "26_knk_scpagwas_overlap_nonlipid8_F1_PD" / "knk_vs_scpagwas_nonlipid8_F1_PD_overlap.tsv",
        ],
    },
    {
        "pair": "lipid8_F1_PD",
        "factor": "lipid8_F1",
        "disease": "PD",
        "label": "lipid8_F1 脳 PD",
        "scp_prefix": "lipid8_F1_MSSM_PD",
        "scp_dir": SCP_BASE / "lipid8_F1_MSSM_PD",
        "knk_summary_paths": [
            BASE / "25_knk_lipid8_F1_PD" / "summary" / "knk_summary_lipid8_F1_PD.csv",
        ],
        "knk_overlap_paths": [
            BASE / "27_knk_scpagwas_overlap_lipid8_F1_PD" / "knk_vs_scpagwas_lipid8_F1_PD_overlap.tsv",
        ],
    },
]


def read_tsv(path: Path) -> pd.DataFrame:
    return pd.read_csv(path, sep="\t", low_memory=False)


def read_csv(path: Path) -> pd.DataFrame:
    return pd.read_csv(path, low_memory=False)


def autosize(ws, max_width: int = 42) -> None:
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells[:200]:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), max_width)


def write_sheet(ws, title: str, subtitle: str, df: pd.DataFrame) -> None:
    ws.delete_rows(1, ws.max_row)
    ws["A1"] = title
    ws["A1"].font = Font(bold=True)
    ws["A2"] = subtitle
    header_row = 4
    for c_idx, col in enumerate(df.columns, start=1):
        ws.cell(row=header_row, column=c_idx, value=col).font = Font(bold=True)
    for r_idx, row in enumerate(df.astype(object).where(pd.notnull(df), "").itertuples(index=False), start=header_row + 1):
        for c_idx, value in enumerate(row, start=1):
            if isinstance(value, float) and (math.isinf(value) or math.isnan(value)):
                value = ""
            ws.cell(row=r_idx, column=c_idx, value=value)
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(df.columns))}{max(4, len(df) + 4)}"
    autosize(ws)


def replace_or_append_sheet(wb, name: str):
    if name in wb.sheetnames:
        idx = wb.sheetnames.index(name)
        ws_old = wb[name]
        wb.remove(ws_old)
        return wb.create_sheet(name, idx)
    return wb.create_sheet(name)


def with_analysis(df: pd.DataFrame, spec: dict) -> pd.DataFrame:
    out = df.copy()
    inserts = [
        ("analysis", spec["label"]),
        ("pair", spec["pair"]),
        ("factor_trait", spec["factor"]),
        ("disease_trait", spec["disease"]),
    ]
    pos = 0
    for col, val in inserts:
        if col not in out.columns:
            out.insert(pos, col, val)
            pos += 1
    return out


def core_candidate_df(spec: dict) -> pd.DataFrame:
    path = BASE / f"17_candidate_gene_integration_{spec['pair']}" / f"{spec['pair']}_cross_evidence_shortlist.tsv"
    df = read_tsv(path)
    return df.loc[df["priority_tier"].isin(["A_high_convergent", "B_multi_source"])].copy()


def core_gene_set(spec: dict) -> set[str]:
    return set(core_candidate_df(spec)["Gene"].dropna().astype(str))


def combine_candidate_core() -> pd.DataFrame:
    dfs = []
    for spec in PAIR_SPECS:
        dfs.append(with_analysis(core_candidate_df(spec), spec))
    return pd.concat(dfs, ignore_index=True)


def filter_gene_level(path: Path, spec: dict, context_type: str, context_label_col: str | None = None) -> pd.DataFrame:
    df = read_tsv(path)
    genes = core_gene_set(spec)
    df = df.loc[df["Gene"].astype(str).isin(genes)].copy()
    df = with_analysis(df, spec)
    if "context_type" not in df.columns:
        df.insert(4, "context_type", context_type)
    if "context_label" not in df.columns:
        if context_label_col is None:
            df.insert(5, "context_label", context_type)
        else:
            df.insert(5, "context_label", df[context_label_col].astype(str))
    return df


def combine_bulk_smr_core() -> pd.DataFrame:
    dfs = []
    for spec in PAIR_SPECS:
        p = BASE / f"11_smr_{spec['pair']}" / "summary" / f"smr_brainmeta_{spec['pair']}_combined.tsv"
        dfs.append(filter_gene_level(p, spec, "BrainMeta_bulk"))
    return pd.concat(dfs, ignore_index=True)


def combine_gtex_smr_core() -> pd.DataFrame:
    dfs = []
    for spec in PAIR_SPECS:
        p = BASE / f"14_smr_gtex_{spec['pair']}" / "summary" / f"smr_gtex_{spec['pair']}_combined.tsv"
        dfs.append(filter_gene_level(p, spec, "GTEx_tissue", "tissue"))
    return pd.concat(dfs, ignore_index=True)


def combine_bryois_smr_core() -> pd.DataFrame:
    dfs = []
    for spec in PAIR_SPECS:
        p = BASE / f"15_smr_bryois_{spec['pair']}" / "summary" / f"smr_bryois_{spec['pair']}_combined.tsv"
        dfs.append(filter_gene_level(p, spec, "Bryois_celltype", "celltype"))
    return pd.concat(dfs, ignore_index=True)


def combine_scpagwas_pathway_curated() -> pd.DataFrame:
    dfs = []
    for spec in PAIR_SPECS:
        trs_dir = spec["scp_dir"] / "Pathway_TRS"
        for file in sorted(trs_dir.glob("Result_*_Pathway_vs_TRS_all.csv")):
            df = read_csv(file)
            # Keep a manuscript-friendly pathway view: top 50 per cell type file.
            sort_col = None
            for candidate in ["adj_p", "FDR", "p.adjust", "pvalue", "p_value", "PCC"]:
                if candidate in df.columns:
                    sort_col = candidate
                    break
            if sort_col is not None:
                ascending = sort_col.lower() != "pcc"
                df = df.sort_values(sort_col, ascending=ascending).head(50).copy()
            else:
                df = df.head(50).copy()
            df = with_analysis(df, spec)
            df["source_file"] = file.name
            dfs.append(df)
    return pd.concat(dfs, ignore_index=True)


def combine_knk_summary_all() -> pd.DataFrame:
    dfs = []
    for spec in PAIR_SPECS:
        for path in spec["knk_summary_paths"]:
            df = with_analysis(read_csv(path), spec)
            df["source_file"] = path.name
            dfs.append(df)
    return pd.concat(dfs, ignore_index=True)


def combine_knk_overlap_all() -> pd.DataFrame:
    dfs = []
    for spec in PAIR_SPECS:
        for path in spec["knk_overlap_paths"]:
            df = read_tsv(path)
            if "scpagwas_match_available" not in df.columns:
                rename_map = {
                    "knk_top20_vs_pericyte_trs_top20_overlap_n": "knk_top20_vs_cell_trs_top20_overlap_n",
                    "knk_top20_vs_pericyte_trs_top20_overlap_ids": "knk_top20_vs_cell_trs_top20_overlap_ids",
                    "knk_top20_vs_pericyte_trs_top20_jaccard": "knk_top20_vs_cell_trs_top20_jaccard",
                    "knk_top50_vs_pericyte_trs_top50_overlap_n": "knk_top50_vs_cell_trs_top50_overlap_n",
                    "knk_top50_vs_pericyte_trs_top50_overlap_ids": "knk_top50_vs_cell_trs_top50_overlap_ids",
                    "knk_top50_vs_pericyte_trs_top50_jaccard": "knk_top50_vs_cell_trs_top50_jaccard",
                }
                df = df.rename(columns=rename_map)
                df["scpagwas_match_available"] = True
                df["scpagwas_reference_cell"] = "pericyte"
                df["remarks"] = ""
            df = with_analysis(df, spec)
            df["source_file"] = path.name
            dfs.append(df)
    return pd.concat(dfs, ignore_index=True)


def build_audit() -> pd.DataFrame:
    rows = [
        {
            "section": "Why v4 is smaller",
            "status": "intentional",
            "notes": (
                "This slim submission workbook keeps manuscript-facing core records only. "
                "Massive gene-level sheets are restricted to A_high_convergent and B_multi_source genes."
            ),
        },
        {
            "section": "Full machine-readable outputs",
            "status": "externalized",
            "notes": (
                "Full SMR/scPagwas/KNK tables remain available as TSV/CSV files in the results directories and should be deposited separately if needed."
            ),
        },
        {
            "section": "Core-gene rule",
            "status": "applied",
            "notes": "Core genes were defined from each pair-specific cross-evidence shortlist using priority_tier in {A_high_convergent, B_multi_source}.",
        },
        {
            "section": "lipid8_F1-PD KNK boundary",
            "status": "retained",
            "notes": "DGKQ and ARL17B were absent from the PD expression matrix and remain explicitly marked as not executable rather than omitted.",
        },
    ]
    return pd.DataFrame(rows)


def build_contents(sheetnames: list[str]) -> pd.DataFrame:
    rows = [{"Order": i + 1, "Sheet name": name} for i, name in enumerate(sheetnames)]
    return pd.DataFrame(rows)


def main() -> None:
    shutil.copy2(V2, OUT)
    wb = load_workbook(OUT)

    additions = {
        "S21a_Candidate_core_genes": (
            "Core candidate genes across three analyses",
            "Only A_high_convergent and B_multi_source records retained for manuscript-facing supplement use.",
            combine_candidate_core(),
        ),
        "S23a_BulkSMR_core_gene_level": (
            "Bulk Brain SMR gene-level evidence (core genes only)",
            "Gene-level BrainMeta SMR rows restricted to core candidate genes.",
            combine_bulk_smr_core(),
        ),
        "S24a_GTExSMR_core_gene_level": (
            "GTEx SMR gene-level evidence (core genes only)",
            "Gene-level GTEx SMR rows restricted to core candidate genes.",
            combine_gtex_smr_core(),
        ),
        "S25a_BryoisSMR_core_gene_level": (
            "Bryois cell-type SMR gene-level evidence (core genes only)",
            "Gene-level Bryois SMR rows restricted to core candidate genes.",
            combine_bryois_smr_core(),
        ),
        "S29a_scPagwas_pathway_curated": (
            "scPagwas pathway results (curated manuscript subset)",
            "Top 50 pathways per scPagwas Pathway_TRS result file for manuscript-facing comparison and KNK overlap interpretation.",
            combine_scpagwas_pathway_curated(),
        ),
        "S31a_KNK_summary_all": (
            "KNK summary across AD and PD analyses",
            "Combined KNK task-level summaries across lipid8_F2-AD, nonlipid8_F1-PD, and lipid8_F1-PD.",
            combine_knk_summary_all(),
        ),
        "S31b_KNK_overlap_all": (
            "KNK versus scPagwas pathway overlap across AD and PD analyses",
            "Combined overlap tables across lipid8_F2-AD, nonlipid8_F1-PD, and lipid8_F1-PD.",
            combine_knk_overlap_all(),
        ),
        "Audit_submission_slim": (
            "Submission workbook slim audit",
            "Notes on why this workbook is smaller and how full machine-readable tables were externalized.",
            build_audit(),
        ),
    }

    for name, (title, subtitle, df) in additions.items():
        ws = replace_or_append_sheet(wb, name)
        write_sheet(ws, title, subtitle, df)

    contents = build_contents(wb.sheetnames)
    write_sheet(wb["Contents"], "Triplet supplementary workbook (slim submission version)", "Core manuscript-facing workbook with full machine-readable outputs externalized to TSV/CSV.", contents)

    wb.save(OUT)
    print(f"Wrote slim workbook to: {OUT}")


if __name__ == "__main__":
    main()

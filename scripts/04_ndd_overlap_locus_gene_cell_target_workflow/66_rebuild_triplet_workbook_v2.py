from __future__ import annotations

import math
import shutil
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


BASE = Path(r"D:\codex\GenomicSEM\metabolic\postgwas_ad_pdlbd\results")
SCP_BASE = Path(r"D:\scPagwas\metabolic_scpagwas2")
TEMPLATE = BASE / "22_supplement_tables_lipid8_F2_AD" / "lipid8_F2_AD_supplementary_tables.xlsx"
OUT_DIR = BASE / "22_supplement_tables_lipid8_F2_AD"
OUT_XLSX = OUT_DIR / "metabolic_factor_triplet_supplementary_tables.xlsx"
OUT_XLSX_V2 = OUT_DIR / "metabolic_factor_triplet_supplementary_tables_v2.xlsx"

PAIR_SPECS = [
    {
        "pair": "lipid8_F2_AD",
        "factor": "lipid8_F2",
        "disease": "AD",
        "label": "lipid8_F2 × AD",
        "traits": ["lipid8_F2", "AD"],
        "scp_prefix": "lipid8_F2_MSSM_AD",
        "scp_dir": BASE / "12_scpagwas2_lipid8_F2_MSSM_AD",
        "atlas_file": "MSSM_AD_20k.rds",
        "atlas_disease": "MSSM_AD",
        "case_status_column": "diagnosis",
        "case_status_value": "AD",
        "atlas_notes": "Downsampled 20k atlas used for scPagwas2 projection",
    },
    {
        "pair": "nonlipid8_F1_PD",
        "factor": "nonlipid8_F1",
        "disease": "PD",
        "label": "nonlipid8_F1 × PD",
        "traits": ["nonlipid8_F1", "PD"],
        "scp_prefix": "nonlipid8_F1_MSSM_PD",
        "scp_dir": SCP_BASE / "nonlipid8_F1_MSSM_PD",
        "atlas_file": "MSSM_PD_20k_from_ALPS_PD.rds",
        "atlas_disease": "MSSM_PD",
        "case_status_column": "PD_status",
        "case_status_value": "PD",
        "atlas_notes": "ALPS-aligned 20k PD atlas used for scPagwas2 projection",
    },
    {
        "pair": "lipid8_F1_PD",
        "factor": "lipid8_F1",
        "disease": "PD",
        "label": "lipid8_F1 × PD",
        "traits": ["lipid8_F1", "PD"],
        "scp_prefix": "lipid8_F1_MSSM_PD",
        "scp_dir": SCP_BASE / "lipid8_F1_MSSM_PD",
        "atlas_file": "MSSM_PD_20k_from_ALPS_PD.rds",
        "atlas_disease": "MSSM_PD",
        "case_status_column": "PD_status",
        "case_status_value": "PD",
        "atlas_notes": "ALPS-aligned 20k PD atlas used for scPagwas2 projection",
    },
]


def read_tsv(path: Path, **kwargs) -> pd.DataFrame:
    return pd.read_csv(path, sep="\t", **kwargs)


def read_csv(path: Path, **kwargs) -> pd.DataFrame:
    return pd.read_csv(path, **kwargs)


def autosize(ws, max_width: int = 42) -> None:
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells[:200]:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), max_width)


def write_sheet(ws, title: str, subtitle: str, df: pd.DataFrame) -> None:
    ws.delete_rows(1, ws.max_row)
    ws["A1"] = title
    ws["A1"].font = Font(bold=True)
    ws["A2"] = subtitle
    header_row = 4
    for c_idx, col in enumerate(df.columns, start=1):
        ws.cell(row=header_row, column=c_idx, value=col).font = Font(bold=True)
    for r_idx, row in enumerate(df.astype(object).where(pd.notnull(df), "").itertuples(index=False), start=header_row + 1):
        for c_idx, value in enumerate(row, start=1):
            if isinstance(value, float) and (math.isinf(value) or math.isnan(value)):
                value = ""
            ws.cell(row=r_idx, column=c_idx, value=value)
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(df.columns))}{max(4, len(df) + 4)}"
    autosize(ws)


def replace_sheet(wb, name: str):
    idx = wb.sheetnames.index(name)
    ws_old = wb[name]
    wb.remove(ws_old)
    return wb.create_sheet(name, idx)


def with_analysis(df: pd.DataFrame, spec: dict) -> pd.DataFrame:
    out = df.copy()
    for pos, (col, val) in enumerate(
        [("analysis", spec["label"]), ("pair", spec["pair"]), ("factor_trait", spec["factor"]), ("disease_trait", spec["disease"])]
    ):
        if col not in out.columns:
            out.insert(pos, col, val)
    return out


def concat_triplet(func) -> pd.DataFrame:
    return pd.concat([func(spec) for spec in PAIR_SPECS], ignore_index=True)


def pair_rg(spec: dict) -> pd.DataFrame:
    rg = read_tsv(BASE / "03_ldsc_metabolic_factors_vs_ndd" / "metabolic_factors_vs_ndd_requested_pairs.tsv")
    sub = rg[(rg["trait1"] == spec["factor"]) & (rg["trait2"] == spec["disease"])].copy()
    return with_analysis(sub, spec)


def pair_rg_mixer(spec: dict) -> pd.DataFrame:
    rg = pair_rg(spec).drop(columns=["analysis", "pair", "factor_trait", "disease_trait"])
    mixer = read_tsv(BASE / f"05_mixer_{spec['pair']}" / f"mixer_{spec['pair']}_summary.tsv")
    out = rg.merge(mixer, on=["trait1", "trait2"], how="left")
    return with_analysis(out, spec)


def pair_pleio_summary(spec: dict) -> pd.DataFrame:
    return with_analysis(read_csv(BASE / f"06_pleiofdr_{spec['pair']}" / f"pleiofdr_{spec['pair']}_summary.csv"), spec)


def pair_pleio_loci(spec: dict) -> pd.DataFrame:
    return with_analysis(read_csv(BASE / f"06_pleiofdr_{spec['pair']}" / f"{spec['pair']}_conjfdr_0.05_loci.csv"), spec)


def pair_coloc(spec: dict) -> pd.DataFrame:
    return with_analysis(read_tsv(BASE / f"09_coloc_{spec['pair']}" / f"coloc_{spec['pair']}_regions.tsv"), spec)


def pair_pwcoco(spec: dict) -> pd.DataFrame:
    priority = BASE / f"10_pwcoco_{spec['pair']}" / f"coloc_pwcoco_{spec['pair']}_priority_regions.tsv"
    fallback = BASE / f"10_pwcoco_{spec['pair']}" / f"pwcoco_{spec['pair']}_best_h4.tsv"
    path = priority if priority.exists() else fallback
    return with_analysis(read_tsv(path), spec)


def locate_fuma_genes(spec: dict) -> Path:
    root = BASE / f"13_fuma_{spec['pair']}"
    candidates = [p for p in root.rglob("genes.txt") if "FUMA_job" in str(p.parent)]
    if not candidates:
        candidates = list(root.rglob("genes.txt"))
    candidates.sort(key=lambda p: (len(p.parts), str(p)))
    return candidates[0]


def pair_fuma(spec: dict) -> pd.DataFrame:
    return with_analysis(read_tsv(locate_fuma_genes(spec)), spec)


def pair_snp_evidence(spec: dict) -> pd.DataFrame:
    return with_analysis(read_tsv(BASE / f"20_snp_evidence_{spec['pair']}" / f"{spec['pair']}_lead_snp_evidence_table.tsv"), spec)


def pair_ctwas_counts(spec: dict) -> pd.DataFrame:
    dfs = [read_tsv(BASE / f"16_ctwas_{spec['pair']}" / "summary" / f"{trait}_counts.tsv") for trait in spec["traits"]]
    return with_analysis(pd.concat(dfs, ignore_index=True), spec)


def pair_candidate_shortlist(spec: dict) -> pd.DataFrame:
    return with_analysis(read_tsv(BASE / f"17_candidate_gene_integration_{spec['pair']}" / f"{spec['pair']}_cross_evidence_shortlist.tsv"), spec)


def pair_candidate_tiers(spec: dict) -> pd.DataFrame:
    return with_analysis(read_tsv(BASE / f"17_candidate_gene_integration_{spec['pair']}" / f"{spec['pair']}_cross_evidence_tier_counts.tsv"), spec)


def pair_brainmeta_summary(spec: dict) -> pd.DataFrame:
    df = read_tsv(BASE / f"11_smr_{spec['pair']}" / "summary" / f"smr_brainmeta_{spec['pair']}_combined.tsv")
    rows = []
    for trait, sub in df.groupby("trait"):
        sub = sub.copy()
        sub["p_SMR_num"] = pd.to_numeric(sub["p_SMR"], errors="coerce")
        sub["p_HEIDI_num"] = pd.to_numeric(sub["p_HEIDI"], errors="coerce")
        best = sub.sort_values("p_SMR_num").iloc[0]
        rows.append(
            {
                "trait": trait,
                "n_rows": len(sub),
                "n_unique_genes": int(sub["Gene"].nunique()),
                "n_smr_p_lt_5e_6": int((sub["p_SMR_num"] < 5e-6).sum()),
                "n_smr_heidi_pass": int(((sub["p_SMR_num"] < 5e-6) & (sub["p_HEIDI_num"] > 0.01)).sum()),
                "best_gene": best["Gene"],
                "best_topSNP": best["topSNP"],
                "best_p_SMR": best["p_SMR"],
                "best_p_HEIDI": best["p_HEIDI"],
            }
        )
    return with_analysis(pd.DataFrame(rows), spec)


def summarize_smr_traitlevel(path: Path, feature_col: str) -> pd.DataFrame:
    df = read_tsv(path)
    df["p_SMR_num"] = pd.to_numeric(df["p_SMR"], errors="coerce")
    df["p_HEIDI_num"] = pd.to_numeric(df["p_HEIDI"], errors="coerce")
    rows = []
    for trait, sub in df.groupby("trait"):
        rows.append(
            {
                "trait": trait,
                "n_rows": len(sub),
                "n_unique_genes": int(sub["Gene"].nunique() if "Gene" in sub.columns else sub["ProbeID"].nunique() if "ProbeID" in sub.columns else pd.Series(dtype=object).nunique()),
                "n_smr_p_lt_5e_6": int((sub["p_SMR_num"] < 5e-6).sum()),
                "n_smr_heidi_pass": int(((sub["p_SMR_num"] < 5e-6) & (sub["p_HEIDI_num"] > 0.01)).sum()),
            }
        )
    return pd.DataFrame(rows)


def pair_gtex_summary(spec: dict) -> pd.DataFrame:
    path = BASE / f"14_smr_gtex_{spec['pair']}" / "summary" / f"smr_gtex_{spec['pair']}_combined.tsv"
    return with_analysis(summarize_smr_traitlevel(path, "tissue"), spec)


def pair_bryois_summary(spec: dict) -> pd.DataFrame:
    path = BASE / f"15_smr_bryois_{spec['pair']}" / "summary" / f"smr_bryois_{spec['pair']}_combined.tsv"
    return with_analysis(summarize_smr_traitlevel(path, "celltype"), spec)


def pair_scpagwas_atlas(spec: dict) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "analysis": spec["scp_prefix"],
                "Disease": spec["atlas_disease"],
                "File": spec["atlas_file"],
                "Total_cells": 20000,
                "Case_status_column": spec["case_status_column"],
                "Case_status_value": spec["case_status_value"],
                "Notes": spec["atlas_notes"],
            }
        ]
    )


def pair_scpagwas_overview(spec: dict) -> pd.DataFrame:
    prefix = spec["scp_prefix"]
    run_dir = spec["scp_dir"]
    merged = read_csv(run_dir / f"{prefix}_Merged_celltype_pvalue_withFDR.csv")
    sig = read_csv(run_dir / f"{prefix}_significant_celltypes.csv")
    path_sig = read_csv(run_dir / f"{prefix}_pathway_summary_sig.csv")
    top = merged.sort_values("celltype_FDR").iloc[0]
    suggestive = merged.loc[pd.to_numeric(merged["celltype_FDR"], errors="coerce") < 0.1, "celltype"].dropna().astype(str).tolist()
    return pd.DataFrame(
        [
            {
                "analysis": prefix,
                "n_sig_celltypes": int(len(sig)),
                "sig_celltypes": ";".join(sig["celltype"].astype(str).tolist()) if len(sig) else "",
                "n_sig_pathways_strict_fdr": int(len(path_sig)),
                "top_celltype_by_fdr": top["celltype"],
                "top_celltype_fdr": top["celltype_FDR"],
                "n_celltypes_tested": int(len(merged)),
                "suggestive_celltypes_fdr_lt_0_1": ";".join(suggestive),
            }
        ]
    )


def pair_scpagwas_celltypes(spec: dict) -> pd.DataFrame:
    prefix = spec["scp_prefix"]
    df = read_csv(spec["scp_dir"] / f"{prefix}_Merged_celltype_pvalue_withFDR.csv")
    df["analysis"] = prefix
    return df[["row_id", "celltype", "pvalue", "celltype_FDR", "analysis"]]


def pair_scpagwas_pathways(spec: dict) -> pd.DataFrame:
    prefix = spec["scp_prefix"]
    run_dir = spec["scp_dir"]
    merged = read_csv(run_dir / f"{prefix}_Merged_celltype_pvalue_withFDR.csv").sort_values("celltype_FDR")
    top_cell = str(merged.iloc[0]["celltype"])
    safe = "".join(ch if ch.isalnum() else "_" for ch in top_cell).strip("_")
    candidate = run_dir / "Pathway_TRS" / f"Result_{safe}_Pathway_vs_TRS_all.csv"
    if not candidate.exists():
        # fallback to first available pathway result
        files = sorted((run_dir / "Pathway_TRS").glob("Result_*_Pathway_vs_TRS_all.csv"))
        if not files:
            return pd.DataFrame(columns=["pathway_id","cell_type","n_genes","n_cells","r","p_value","FDR","pathway_name","abs_r","analysis"])
        candidate = files[0]
    df = read_csv(candidate)
    df["analysis"] = prefix
    return df[["pathway_id", "cell_type", "n_genes", "n_cells", "r", "p_value", "FDR", "pathway_name", "abs_r", "analysis"]]


def pair_pcc_curated(spec: dict) -> pd.DataFrame:
    prefix = spec["scp_prefix"]
    pcc = read_csv(spec["scp_dir"] / f"{prefix}_gene_PCC.csv").rename(columns={"Unnamed: 0": "gene"})
    shortlist = read_tsv(BASE / f"17_candidate_gene_integration_{spec['pair']}" / f"{spec['pair']}_cross_evidence_shortlist.tsv")
    genes = shortlist["Gene"].dropna().astype(str).unique().tolist()
    out = pcc[pcc["gene"].astype(str).isin(genes)].copy()
    out["analysis"] = prefix
    return out[["gene", "PCC", "pvalue", "adj_logp", "adj_pvalue", "weight_pcc", "analysis"]]


def build_audit() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {"section": "S24/S25 trait-level summaries", "status": "fixed", "notes": "Rebuilt from combined SMR tables for all three analyses to match a single trait-level schema."},
            {"section": "PD scPagwas2 pathways", "status": "fixed", "notes": "Both PD lines were post-processed with the same pathway/TRS summarizer used for F2-AD."},
            {"section": "S26-S30 scPagwas2 integration", "status": "fixed", "notes": "Workbook now carries atlas metadata, overview, celltypes, pathways, and PCC in a three-analysis format."},
            {"section": "Single-cell / KNK downstream", "status": "pending for PD", "notes": "F2-AD KNK remains; PD KNK not yet run."},
        ]
    )


def build_contents(sheetnames: list[str]) -> pd.DataFrame:
    rows = []
    for idx, name in enumerate(sheetnames, start=1):
        rows.append({"Supplementary Table": f"S{idx}" if name.startswith("S") else "", "Sheet name": name})
    return pd.DataFrame(rows)


def main() -> None:
    shutil.copy2(TEMPLATE, OUT_XLSX)
    wb = load_workbook(OUT_XLSX)

    replacements = {
        "S3_rg_panel": ("Cross-disease rg panel", "Three focal factor-disease pairs are shown together for direct comparison.", concat_triplet(pair_rg)),
        "S13_CrossDisease_LDSC_MiXeR": ("Cross-disease LDSC + MiXeR", "Merged LDSC pairwise rg and MiXeR summaries across the three focal pairs.", concat_triplet(pair_rg_mixer)),
        "S14_PleioFDR_summary": ("pleioFDR summary", "One-line pleioFDR/conjFDR summaries across the three focal pairs.", concat_triplet(pair_pleio_summary)),
        "S15_PleioFDR_loci": ("conjFDR loci", "Sentinel loci passing conjFDR < 0.05 across the three focal pairs.", concat_triplet(pair_pleio_loci)),
        "S16_Coloc_summary": ("coloc regional summary", "Regional coloc summaries across the three focal pairs.", concat_triplet(pair_coloc)),
        "S17_PWCoCo_best": ("PWCoCo prioritized regions", "Priority/shared-signal PWCoCo results across the three focal pairs.", concat_triplet(pair_pwcoco)),
        "S18_FUMA_mapping_summary": ("FUMA mapping summary", "FUMA gene mapping outputs across the three focal pairs.", concat_triplet(pair_fuma)),
        "S19_SNP_evidence": ("Lead SNP evidence", "Lead-SNP evidence tables across the three focal pairs, combining conjFDR, coloc, PWCoCo, and FUMA annotations.", concat_triplet(pair_snp_evidence)),
        "S20_cTWAS_overview": ("cTWAS overview", "Trait-level cTWAS summary counts across the three focal pairs.", concat_triplet(pair_ctwas_counts)),
        "S21_Candidate_master": ("Candidate shortlist", "High-priority cross-evidence candidate shortlist across the three focal pairs, using the same logic as the original F2-AD workbook.", concat_triplet(pair_candidate_shortlist)),
        "S22_Candidate_tier_summary": ("Candidate tier summary", "Priority tier counts across the three focal pairs.", concat_triplet(pair_candidate_tiers)),
        "S23_BulkBrain_SMR_summary": ("Bulk-brain SMR summary", "BrainMeta SMR summaries across the three focal pairs.", concat_triplet(pair_brainmeta_summary)),
        "S24_GTEx_SMR_summary": ("GTEx SMR summary", "Trait-level GTEx SMR summaries across the three focal pairs using a unified schema.", concat_triplet(pair_gtex_summary)),
        "S25_CellType_SMR_summary": ("Cell-type SMR summary", "Trait-level Bryois cell-type SMR summaries across the three focal pairs using a unified schema.", concat_triplet(pair_bryois_summary)),
        "S26_scPagwas_atlas": ("Single-nucleus atlas metadata", "Atlas metadata for the three scPagwas2 analyses.", concat_triplet(pair_scpagwas_atlas)),
        "S27_scPagwas_overview": ("scPagwas2 overview", "Overall scPagwas2 summaries across the three focal analyses.", pd.concat([pair_scpagwas_overview(spec) for spec in PAIR_SPECS], ignore_index=True)),
        "S28_F2_AD_celltypes": ("scPagwas2 cell-type results", "Cell-type enrichment results across the three focal analyses.", pd.concat([pair_scpagwas_celltypes(spec) for spec in PAIR_SPECS], ignore_index=True)),
        "S29_F2_AD_pathways": ("scPagwas2 pathway-TRS results", "Key cell-type pathway-TRS result tables across the three focal analyses.", pd.concat([pair_scpagwas_pathways(spec) for spec in PAIR_SPECS], ignore_index=True)),
        "S30_PCC_curated": ("Curated gene PCC results", "Curated gene-level PCC results for shortlisted candidate genes across the three focal analyses.", pd.concat([pair_pcc_curated(spec) for spec in PAIR_SPECS], ignore_index=True)),
    }

    for name, (title, subtitle, df) in replacements.items():
        ws = replace_sheet(wb, name)
        write_sheet(ws, title, subtitle, df)

    if "Audit_triplet" in wb.sheetnames:
        wb.remove(wb["Audit_triplet"])
    ws_audit = wb.create_sheet("Audit_triplet", 1)
    write_sheet(ws_audit, "Triplet integration audit", "Audit of rebuilt triplet workbook after harmonizing SMR summaries and PD scPagwas2 outputs.", build_audit())

    write_sheet(wb["Contents"], "Contents", "Triplet-comparison supplementary workbook with harmonized region/gene/single-cell sheets.", build_contents(wb.sheetnames))
    wb.save(OUT_XLSX_V2)
    try:
        wb.save(OUT_XLSX)
    except OSError:
        pass


if __name__ == "__main__":
    main()

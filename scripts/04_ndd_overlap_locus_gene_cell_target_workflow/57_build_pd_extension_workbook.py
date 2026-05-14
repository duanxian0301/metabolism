from __future__ import annotations

import math
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


BASE = Path(r"D:\codex\GenomicSEM\metabolic\postgwas_ad_pdlbd\results")
OUT_DIR = BASE / "23_pd_factor_extensions"
OUT_DIR.mkdir(parents=True, exist_ok=True)
OUT_XLSX = OUT_DIR / "pd_factor_extensions_supplementary_tables.xlsx"


def read_tsv(path: Path, **kwargs) -> pd.DataFrame:
    return pd.read_csv(path, sep="\t", **kwargs)


def read_csv(path: Path, **kwargs) -> pd.DataFrame:
    return pd.read_csv(path, **kwargs)


def autosize(ws, max_width: int = 42) -> None:
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells[:200]:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), max_width)


def add_sheet(wb: Workbook, name: str, title: str, subtitle: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(name[:31])
    ws["A1"] = title
    ws["A1"].font = Font(bold=True)
    ws["A2"] = subtitle
    header_row = 4
    for c_idx, col in enumerate(df.columns, start=1):
        ws.cell(row=header_row, column=c_idx, value=col).font = Font(bold=True)
    for r_idx, row in enumerate(df.astype(object).where(pd.notnull(df), "").itertuples(index=False), start=header_row + 1):
        for c_idx, value in enumerate(row, start=1):
            if isinstance(value, float) and (math.isinf(value) or math.isnan(value)):
                value = ""
            ws.cell(row=r_idx, column=c_idx, value=value)
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(df.columns))}{max(4, len(df) + 4)}"
    autosize(ws)


def build_contents(rows: list[tuple[str, str, str]]) -> pd.DataFrame:
    return pd.DataFrame(rows, columns=["Sheet", "Topic", "Description"])


def pair_rg_mixer(pair: str, factor: str, disease: str) -> pd.DataFrame:
    rg = read_tsv(BASE / "03_ldsc_metabolic_factors_vs_ndd" / "metabolic_factors_vs_ndd_requested_pairs.tsv")
    mixer = read_tsv(BASE / f"05_mixer_{pair}" / f"mixer_{pair}_summary.tsv")
    sub = rg[(rg["trait1"] == factor) & (rg["trait2"] == disease)].copy()
    return sub.merge(mixer, on=["trait1", "trait2"], how="left")


def pair_brainmeta_summary(pair: str) -> pd.DataFrame:
    df = read_tsv(BASE / f"11_smr_{pair}" / "summary" / f"smr_brainmeta_{pair}_combined.tsv")
    rows = []
    for trait, sub in df.groupby("trait"):
        best = sub.sort_values("p_SMR").iloc[0]
        rows.append(
            {
                "trait": trait,
                "panel": "BrainMeta",
                "context": "bulk_brain",
                "n_rows": len(sub),
                "n_hits_p_SMR_lt_1e4": int((pd.to_numeric(sub["p_SMR"], errors="coerce") < 1e-4).sum()),
                "n_hits_supported": int(((pd.to_numeric(sub["p_SMR"], errors="coerce") < 1e-4) & pd.to_numeric(sub["p_HEIDI"], errors="coerce").gt(0.01)).sum()),
                "best_gene": best["Gene"],
                "best_probe": best.get("probeID", ""),
                "best_topSNP": best["topSNP"],
                "best_p_SMR": best["p_SMR"],
                "best_p_HEIDI": best["p_HEIDI"],
            }
        )
    return pd.DataFrame(rows)


def build_audit() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "analysis_step": "LDSC pairwise rg",
                "nonlipid8_F1_PD": "ready",
                "lipid8_F1_PD": "ready",
                "notes": "Significant PD-linked factor pairs retained for extension",
            },
            {
                "analysis_step": "MiXeR",
                "nonlipid8_F1_PD": "ready",
                "lipid8_F1_PD": "ready",
                "notes": "pair-specific MiXeR summaries available",
            },
            {
                "analysis_step": "pleioFDR / conjFDR loci",
                "nonlipid8_F1_PD": "ready",
                "lipid8_F1_PD": "ready",
                "notes": "local 06_pleiofdr_* directories staged from pleioFDR outputs",
            },
            {
                "analysis_step": "coloc / PWCoCo",
                "nonlipid8_F1_PD": "ready",
                "lipid8_F1_PD": "ready",
                "notes": "priority region tables available",
            },
            {
                "analysis_step": "FUMA",
                "nonlipid8_F1_PD": "ready",
                "lipid8_F1_PD": "ready",
                "notes": "website outputs mapped to pair-specific directories",
            },
            {
                "analysis_step": "BrainMeta SMR",
                "nonlipid8_F1_PD": "ready",
                "lipid8_F1_PD": "ready",
                "notes": "factor-side frequency-check issue resolved using F2-AD logic",
            },
            {
                "analysis_step": "GTEx SMR",
                "nonlipid8_F1_PD": "ready",
                "lipid8_F1_PD": "ready",
                "notes": "26/26 tissues completed for both pairs",
            },
            {
                "analysis_step": "Bryois SMR",
                "nonlipid8_F1_PD": "ready",
                "lipid8_F1_PD": "ready",
                "notes": "352/352 celltype-chromosome tasks completed for both pairs",
            },
            {
                "analysis_step": "cTWAS",
                "nonlipid8_F1_PD": "ready",
                "lipid8_F1_PD": "ready",
                "notes": "local summary directories exported from WSL outputs",
            },
            {
                "analysis_step": "Candidate gene integration",
                "nonlipid8_F1_PD": "ready",
                "lipid8_F1_PD": "ready",
                "notes": "same logic and columns as F2-AD cross-evidence table",
            },
            {
                "analysis_step": "scPagwas2",
                "nonlipid8_F1_PD": "pending",
                "lipid8_F1_PD": "pending",
                "notes": "next planned step; not yet run for PD lines",
            },
            {
                "analysis_step": "virtual KNK",
                "nonlipid8_F1_PD": "pending",
                "lipid8_F1_PD": "pending",
                "notes": "to be decided after scPagwas2 cell prioritization",
            },
        ]
    )


def main() -> None:
    wb = Workbook()
    ws0 = wb[wb.sheetnames[0]]
    ws0.title = "Contents"

    contents_rows: list[tuple[str, str, str]] = []

    audit = build_audit()
    contents_rows.append(("Audit", "Project audit", "Ready vs pending outputs for the two PD extension lines"))

    add_sheet(wb, "Audit", "PD extension audit", "This sheet records which downstream layers are already complete and which remain pending.", audit)

    pair_specs = [
        ("nonlipid8_F1_PD", "nonlipid8_F1", "PD"),
        ("lipid8_F1_PD", "lipid8_F1", "PD"),
    ]

    for pair, factor, disease in pair_specs:
        tag = "nF1PD" if pair.startswith("nonlipid8") else "lF1PD"

        rg_mix = pair_rg_mixer(pair, factor, disease)
        add_sheet(wb, f"{tag}_rg_MiXeR", f"{pair} LDSC + MiXeR summary", "Cross-trait LDSC result merged with MiXeR summary.", rg_mix)
        contents_rows.append((f"{tag}_rg_MiXeR", pair, "LDSC requested pair result merged with MiXeR summary"))

        pleio_summary = read_csv(BASE / f"06_pleiofdr_{pair}" / f"pleiofdr_{pair}_summary.csv")
        add_sheet(wb, f"{tag}_pleio", f"{pair} pleioFDR summary", "One-line summary of conjFDR discovery output.", pleio_summary)
        contents_rows.append((f"{tag}_pleio", pair, "pleioFDR summary"))

        loci = read_csv(BASE / f"06_pleiofdr_{pair}" / f"{pair}_conjfdr_0.05_loci.csv")
        add_sheet(wb, f"{tag}_loci", f"{pair} conjFDR loci", "Sentinel loci passing conjFDR < 0.05.", loci)
        contents_rows.append((f"{tag}_loci", pair, "conjFDR loci"))

        coloc = read_tsv(BASE / f"09_coloc_{pair}" / f"coloc_{pair}_regions.tsv")
        add_sheet(wb, f"{tag}_coloc", f"{pair} coloc regions", "Regional coloc summary for the pair.", coloc)
        contents_rows.append((f"{tag}_coloc", pair, "coloc regional summary"))

        pw = read_tsv(BASE / f"10_pwcoco_{pair}" / f"coloc_pwcoco_{pair}_priority_regions.tsv")
        add_sheet(wb, f"{tag}_PWCoCo", f"{pair} PWCoCo priority regions", "Integrated PWCoCo / coloc priority region table.", pw)
        contents_rows.append((f"{tag}_PWCoCo", pair, "PWCoCo priority regions"))

        fuma = read_tsv(BASE / f"13_fuma_{pair}" / [p.name for p in (BASE / f"13_fuma_{pair}").iterdir() if p.is_dir() and p.name.startswith("FUMA_job")][0] / "genes.txt")
        add_sheet(wb, f"{tag}_FUMA", f"{pair} FUMA mapped genes", "FUMA genes.txt output used for downstream integration.", fuma)
        contents_rows.append((f"{tag}_FUMA", pair, "FUMA mapped genes"))

        brainmeta = pair_brainmeta_summary(pair)
        add_sheet(wb, f"{tag}_bulkSMR", f"{pair} BrainMeta SMR summary", "Bulk-brain SMR summary derived from combined BrainMeta outputs.", brainmeta)
        contents_rows.append((f"{tag}_bulkSMR", pair, "BrainMeta SMR summary"))

        gtex_counts = read_tsv(BASE / f"14_smr_gtex_{pair}" / "summary" / f"smr_gtex_{pair}_counts.tsv")
        add_sheet(wb, f"{tag}_GTExSMR", f"{pair} GTEx SMR counts", "GTEx tissue-level SMR count summary.", gtex_counts)
        contents_rows.append((f"{tag}_GTExSMR", pair, "GTEx SMR counts"))

        bryois_counts = read_tsv(BASE / f"15_smr_bryois_{pair}" / "summary" / f"smr_bryois_{pair}_counts.tsv")
        add_sheet(wb, f"{tag}_BryoisSMR", f"{pair} Bryois SMR counts", "Bryois cell-type SMR count summary.", bryois_counts)
        contents_rows.append((f"{tag}_BryoisSMR", pair, "Bryois SMR counts"))

        ctwas_counts = pd.concat(
            [
                read_tsv(BASE / f"16_ctwas_{pair}" / "summary" / f"{factor}_counts.tsv"),
                read_tsv(BASE / f"16_ctwas_{pair}" / "summary" / f"{disease}_counts.tsv"),
            ],
            ignore_index=True,
        )
        add_sheet(wb, f"{tag}_cTWAS", f"{pair} cTWAS counts", "Trait-level cTWAS run summary and object counts.", ctwas_counts)
        contents_rows.append((f"{tag}_cTWAS", pair, "cTWAS counts"))

        cand_short = read_tsv(BASE / f"17_candidate_gene_integration_{pair}" / f"{pair}_cross_evidence_shortlist.tsv")
        add_sheet(wb, f"{tag}_candShort", f"{pair} candidate gene shortlist", "High-priority cross-evidence shortlist using the same logic and columns as F2-AD.", cand_short)
        contents_rows.append((f"{tag}_candShort", pair, "Candidate shortlist"))

        cand_master = read_tsv(BASE / f"17_candidate_gene_integration_{pair}" / f"{pair}_cross_evidence_master.tsv")
        add_sheet(wb, f"{tag}_candMaster", f"{pair} candidate gene master table", "Full cross-evidence candidate gene table using the same logic and columns as F2-AD.", cand_master)
        contents_rows.append((f"{tag}_candMaster", pair, "Candidate master table"))

        cand_tiers = read_tsv(BASE / f"17_candidate_gene_integration_{pair}" / f"{pair}_cross_evidence_tier_counts.tsv")
        add_sheet(wb, f"{tag}_candTier", f"{pair} candidate tier counts", "Priority tier counts for the pair-specific cross-evidence table.", cand_tiers)
        contents_rows.append((f"{tag}_candTier", pair, "Candidate tier counts"))

    contents_df = build_contents(contents_rows)
    ws = wb["Contents"]
    ws["A1"] = "Contents"
    ws["A1"].font = Font(bold=True)
    ws["A2"] = "PD extension supplementary workbook for nonlipid8_F1×PD and lipid8_F1×PD."
    for c_idx, col in enumerate(contents_df.columns, start=1):
        ws.cell(row=4, column=c_idx, value=col).font = Font(bold=True)
    for r_idx, row in enumerate(contents_df.itertuples(index=False), start=5):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(contents_df.columns))}{len(contents_df)+4}"
    autosize(ws)

    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}")


if __name__ == "__main__":
    main()

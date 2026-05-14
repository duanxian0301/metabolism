from __future__ import annotations

import math
import shutil
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


BASE = Path(r"D:\codex\GenomicSEM\metabolic\postgwas_ad_pdlbd\results")
SCP_BASE = Path(r"D:\scPagwas\metabolic_scpagwas2")
V2 = BASE / "22_supplement_tables_lipid8_F2_AD" / "metabolic_factor_triplet_supplementary_tables_v2.xlsx"
OUT = BASE / "22_supplement_tables_lipid8_F2_AD" / "metabolic_factor_triplet_supplementary_tables_v3_submission.xlsx"


PAIR_SPECS = [
    {
        "pair": "lipid8_F2_AD",
        "factor": "lipid8_F2",
        "disease": "AD",
        "label": "lipid8_F2 × AD",
        "scp_prefix": "lipid8_F2_MSSM_AD",
        "scp_dir": SCP_BASE / "lipid8_F2_MSSM_AD",
        "knk_summary_paths": [
            BASE / "19_knk_lipid8_F2_AD_core4" / "summary" / "knk_summary_core4_shard1of2.csv",
            BASE / "19_knk_lipid8_F2_AD_core4" / "summary" / "knk_summary_core4_shard2of2.csv",
        ],
        "knk_overlap_paths": [
            BASE / "21_knk_scpagwas_overlap_lipid8_F2_AD" / "knk_vs_scpagwas_pericyte_trs_overlap.tsv",
        ],
    },
    {
        "pair": "nonlipid8_F1_PD",
        "factor": "nonlipid8_F1",
        "disease": "PD",
        "label": "nonlipid8_F1 × PD",
        "scp_prefix": "nonlipid8_F1_MSSM_PD",
        "scp_dir": SCP_BASE / "nonlipid8_F1_MSSM_PD",
        "knk_summary_paths": [
            BASE / "24_knk_nonlipid8_F1_PD" / "summary" / "knk_summary_nonlipid8_F1_PD.csv",
        ],
        "knk_overlap_paths": [
            BASE / "26_knk_scpagwas_overlap_nonlipid8_F1_PD" / "knk_vs_scpagwas_nonlipid8_F1_PD_overlap.tsv",
        ],
    },
    {
        "pair": "lipid8_F1_PD",
        "factor": "lipid8_F1",
        "disease": "PD",
        "label": "lipid8_F1 × PD",
        "scp_prefix": "lipid8_F1_MSSM_PD",
        "scp_dir": SCP_BASE / "lipid8_F1_MSSM_PD",
        "knk_summary_paths": [
            BASE / "25_knk_lipid8_F1_PD" / "summary" / "knk_summary_lipid8_F1_PD.csv",
        ],
        "knk_overlap_paths": [
            BASE / "27_knk_scpagwas_overlap_lipid8_F1_PD" / "knk_vs_scpagwas_lipid8_F1_PD_overlap.tsv",
        ],
    },
]


def read_tsv(path: Path) -> pd.DataFrame:
    return pd.read_csv(path, sep="\t", low_memory=False)


def read_csv(path: Path) -> pd.DataFrame:
    return pd.read_csv(path, low_memory=False)


def autosize(ws, max_width: int = 42) -> None:
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells[:200]:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), max_width)


def write_sheet(ws, title: str, subtitle: str, df: pd.DataFrame) -> None:
    ws.delete_rows(1, ws.max_row)
    ws["A1"] = title
    ws["A1"].font = Font(bold=True)
    ws["A2"] = subtitle
    header_row = 4
    for c_idx, col in enumerate(df.columns, start=1):
        ws.cell(row=header_row, column=c_idx, value=col).font = Font(bold=True)
    for r_idx, row in enumerate(df.astype(object).where(pd.notnull(df), "").itertuples(index=False), start=header_row + 1):
        for c_idx, value in enumerate(row, start=1):
            if isinstance(value, float) and (math.isinf(value) or math.isnan(value)):
                value = ""
            ws.cell(row=r_idx, column=c_idx, value=value)
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(df.columns))}{max(4, len(df) + 4)}"
    autosize(ws)


def replace_or_append_sheet(wb, name: str):
    if name in wb.sheetnames:
        idx = wb.sheetnames.index(name)
        ws_old = wb[name]
        wb.remove(ws_old)
        return wb.create_sheet(name, idx)
    return wb.create_sheet(name)


def with_analysis(df: pd.DataFrame, spec: dict) -> pd.DataFrame:
    out = df.copy()
    inserts = [
        ("analysis", spec["label"]),
        ("pair", spec["pair"]),
        ("factor_trait", spec["factor"]),
        ("disease_trait", spec["disease"]),
    ]
    pos = 0
    for col, val in inserts:
        if col not in out.columns:
            out.insert(pos, col, val)
            pos += 1
    return out


def combine_candidate_master_full() -> pd.DataFrame:
    dfs = []
    for spec in PAIR_SPECS:
        p = BASE / f"17_candidate_gene_integration_{spec['pair']}" / f"{spec['pair']}_cross_evidence_master.tsv"
        dfs.append(with_analysis(read_tsv(p), spec))
    return pd.concat(dfs, ignore_index=True)


def combine_bulk_smr_gene_level() -> pd.DataFrame:
    dfs = []
    for spec in PAIR_SPECS:
        p = BASE / f"11_smr_{spec['pair']}" / "summary" / f"smr_brainmeta_{spec['pair']}_combined.tsv"
        df = with_analysis(read_tsv(p), spec)
        if "context_type" not in df.columns:
            df.insert(4, "context_type", "BrainMeta_bulk")
        if "context_label" not in df.columns:
            df.insert(5, "context_label", "BrainMeta_bulk")
        dfs.append(df)
    return pd.concat(dfs, ignore_index=True)


def combine_gtex_smr_gene_level() -> pd.DataFrame:
    dfs = []
    for spec in PAIR_SPECS:
        p = BASE / f"14_smr_gtex_{spec['pair']}" / "summary" / f"smr_gtex_{spec['pair']}_combined.tsv"
        df = with_analysis(read_tsv(p), spec)
        if "context_type" not in df.columns:
            df.insert(4, "context_type", "GTEx_tissue")
        if "context_label" not in df.columns:
            df.insert(5, "context_label", df["tissue"].astype(str))
        dfs.append(df)
    return pd.concat(dfs, ignore_index=True)


def combine_bryois_smr_gene_level() -> pd.DataFrame:
    dfs = []
    for spec in PAIR_SPECS:
        p = BASE / f"15_smr_bryois_{spec['pair']}" / "summary" / f"smr_bryois_{spec['pair']}_combined.tsv"
        df = with_analysis(read_tsv(p), spec)
        if "context_type" not in df.columns:
            df.insert(4, "context_type", "Bryois_celltype")
        if "context_label" not in df.columns:
            df.insert(5, "context_label", df["celltype"].astype(str))
        dfs.append(df)
    return pd.concat(dfs, ignore_index=True)


def combine_scpagwas_pathway_all() -> pd.DataFrame:
    dfs = []
    for spec in PAIR_SPECS:
        trs_dir = spec["scp_dir"] / "Pathway_TRS"
        for file in sorted(trs_dir.glob("Result_*_Pathway_vs_TRS_all.csv")):
            df = read_csv(file)
            df = with_analysis(df, spec)
            df["source_file"] = file.name
            dfs.append(df)
    return pd.concat(dfs, ignore_index=True)


def combine_knk_summary_all() -> pd.DataFrame:
    dfs = []
    for spec in PAIR_SPECS:
        for path in spec["knk_summary_paths"]:
            df = with_analysis(read_csv(path), spec)
            df["source_file"] = path.name
            dfs.append(df)
    return pd.concat(dfs, ignore_index=True)


def combine_knk_overlap_all() -> pd.DataFrame:
    dfs = []
    for spec in PAIR_SPECS:
        for path in spec["knk_overlap_paths"]:
            df = read_tsv(path)
            if "scpagwas_match_available" not in df.columns:
                # Harmonize older F2-AD overlap schema.
                rename_map = {
                    "knk_top20_vs_pericyte_trs_top20_overlap_n": "knk_top20_vs_cell_trs_top20_overlap_n",
                    "knk_top20_vs_pericyte_trs_top20_overlap_ids": "knk_top20_vs_cell_trs_top20_overlap_ids",
                    "knk_top20_vs_pericyte_trs_top20_jaccard": "knk_top20_vs_cell_trs_top20_jaccard",
                    "knk_top50_vs_pericyte_trs_top50_overlap_n": "knk_top50_vs_cell_trs_top50_overlap_n",
                    "knk_top50_vs_pericyte_trs_top50_overlap_ids": "knk_top50_vs_cell_trs_top50_overlap_ids",
                    "knk_top50_vs_pericyte_trs_top50_jaccard": "knk_top50_vs_cell_trs_top50_jaccard",
                }
                df = df.rename(columns=rename_map)
                df["scpagwas_match_available"] = True
                df["scpagwas_reference_cell"] = "pericyte"
                df["remarks"] = ""
            df = with_analysis(df, spec)
            df["source_file"] = path.name
            dfs.append(df)
    return pd.concat(dfs, ignore_index=True)


def build_submission_audit() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "section": "SMR detailed sheets",
                "status": "added",
                "notes": "Gene-level BrainMeta, GTEx, and Bryois SMR sheets appended for all three analyses.",
            },
            {
                "section": "scPagwas pathway all",
                "status": "added",
                "notes": "Stacked all Pathway_TRS Result_*_all.csv files across AD and both PD lines for pathway-level supplement use.",
            },
            {
                "section": "KNK PD integration",
                "status": "added",
                "notes": "Added nonlipid8_F1-PD and lipid8_F1-PD KNK summaries and scPagwas overlap outputs alongside the F2-AD KNK results.",
            },
            {
                "section": "Real data boundary",
                "status": "retained",
                "notes": "In lipid8_F1-PD KNK, DGKQ and ARL17B were absent from the PD expression matrix; those tasks remain explicitly marked as not executable, not missing.",
            },
        ]
    )


def build_contents(sheetnames: list[str]) -> pd.DataFrame:
    rows = []
    for idx, name in enumerate(sheetnames, start=1):
        rows.append({"Order": idx, "Sheet name": name})
    return pd.DataFrame(rows)


def main() -> None:
    shutil.copy2(V2, OUT)
    wb = load_workbook(OUT)

    additions = {
        "S21a_Candidate_master_full": (
            "Candidate master full",
            "Full cross-evidence candidate master tables across the three focal analyses.",
            combine_candidate_master_full(),
        ),
        "S23a_BulkSMR_gene_level": (
            "Bulk-brain SMR gene-level",
            "Gene-level BrainMeta SMR tables across the three focal analyses.",
            combine_bulk_smr_gene_level(),
        ),
        "S24a_GTExSMR_gene_level": (
            "GTEx SMR gene-level",
            "Gene-level GTEx SMR tables across the three focal analyses.",
            combine_gtex_smr_gene_level(),
        ),
        "S25a_BryoisSMR_gene_level": (
            "Bryois cell-type SMR gene-level",
            "Gene-level Bryois cell-type SMR tables across the three focal analyses.",
            combine_bryois_smr_gene_level(),
        ),
        "S29a_scPagwas_pathway_all": (
            "scPagwas2 pathway all",
            "All scPagwas2 Pathway_TRS per-cell-type pathway association tables across the three focal analyses.",
            combine_scpagwas_pathway_all(),
        ),
        "S31a_KNK_summary_all": (
            "KNK summary all",
            "Combined KNK summaries across lipid8_F2-AD, nonlipid8_F1-PD, and lipid8_F1-PD.",
            combine_knk_summary_all(),
        ),
        "S31b_KNK_overlap_all": (
            "KNK-scPagwas overlap all",
            "Combined KNK versus scPagwas pathway overlap tables across the three focal analyses.",
            combine_knk_overlap_all(),
        ),
        "Audit_submission": (
            "Submission audit",
            "Audit of added detailed sheets for the submission-ready triplet supplementary workbook.",
            build_submission_audit(),
        ),
    }

    for name, (title, subtitle, df) in additions.items():
        ws = replace_or_append_sheet(wb, name)
        write_sheet(ws, title, subtitle, df)

    write_sheet(
        wb["Contents"],
        "Contents",
        "Submission-ready triplet supplementary workbook with overview and detailed sheets.",
        build_contents(wb.sheetnames),
    )

    wb.save(OUT)


if __name__ == "__main__":
    main()

from __future__ import annotations

import math
import shutil
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


BASE = Path(r"D:\codex\GenomicSEM\metabolic\postgwas_ad_pdlbd\results")
TEMPLATE = BASE / "22_supplement_tables_lipid8_F2_AD" / "lipid8_F2_AD_supplementary_tables.xlsx"
OUT_XLSX = BASE / "22_supplement_tables_lipid8_F2_AD" / "metabolic_factor_triplet_supplementary_tables.xlsx"

PAIR_SPECS = [
    {
        "pair": "lipid8_F2_AD",
        "factor": "lipid8_F2",
        "disease": "AD",
        "label": "lipid8_F2 × AD",
        "traits": ["lipid8_F2", "AD"],
    },
    {
        "pair": "nonlipid8_F1_PD",
        "factor": "nonlipid8_F1",
        "disease": "PD",
        "label": "nonlipid8_F1 × PD",
        "traits": ["nonlipid8_F1", "PD"],
    },
    {
        "pair": "lipid8_F1_PD",
        "factor": "lipid8_F1",
        "disease": "PD",
        "label": "lipid8_F1 × PD",
        "traits": ["lipid8_F1", "PD"],
    },
]


def read_tsv(path: Path, **kwargs) -> pd.DataFrame:
    return pd.read_csv(path, sep="\t", **kwargs)


def read_csv(path: Path, **kwargs) -> pd.DataFrame:
    return pd.read_csv(path, **kwargs)


def autosize(ws, max_width: int = 42) -> None:
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells[:200]:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), max_width)


def write_sheet(ws, title: str, subtitle: str, df: pd.DataFrame) -> None:
    ws.delete_rows(1, ws.max_row)
    ws["A1"] = title
    ws["A1"].font = Font(bold=True)
    ws["A2"] = subtitle
    header_row = 4
    for c_idx, col in enumerate(df.columns, start=1):
        ws.cell(row=header_row, column=c_idx, value=col).font = Font(bold=True)
    for r_idx, row in enumerate(df.astype(object).where(pd.notnull(df), "").itertuples(index=False), start=header_row + 1):
        for c_idx, value in enumerate(row, start=1):
            if isinstance(value, float) and (math.isinf(value) or math.isnan(value)):
                value = ""
            ws.cell(row=r_idx, column=c_idx, value=value)
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(df.columns))}{max(4, len(df) + 4)}"
    autosize(ws)


def with_analysis(df: pd.DataFrame, spec: dict, columns_first: list[str] | None = None) -> pd.DataFrame:
    out = df.copy()
    insert_vals = [
        ("analysis", spec["label"]),
        ("pair", spec["pair"]),
        ("factor_trait", spec["factor"]),
        ("disease_trait", spec["disease"]),
    ]
    pos = 0
    for col, val in insert_vals:
        if col in out.columns:
            continue
        out.insert(pos, col, val)
        pos += 1
    if columns_first:
        keep = [c for c in columns_first if c in out.columns]
        rest = [c for c in out.columns if c not in keep]
        out = out[keep + rest]
    return out


def pair_rg_mixer(spec: dict) -> pd.DataFrame:
    rg = read_tsv(BASE / "03_ldsc_metabolic_factors_vs_ndd" / "metabolic_factors_vs_ndd_requested_pairs.tsv")
    mixer = read_tsv(BASE / f"05_mixer_{spec['pair']}" / f"mixer_{spec['pair']}_summary.tsv")
    sub = rg[(rg["trait1"] == spec["factor"]) & (rg["trait2"] == spec["disease"])].copy()
    out = sub.merge(mixer, on=["trait1", "trait2"], how="left")
    return with_analysis(out, spec)


def pair_rg_panel(spec: dict) -> pd.DataFrame:
    rg = read_tsv(BASE / "03_ldsc_metabolic_factors_vs_ndd" / "metabolic_factors_vs_ndd_requested_pairs.tsv")
    sub = rg[(rg["trait1"] == spec["factor"]) & (rg["trait2"] == spec["disease"])].copy()
    return with_analysis(sub, spec)


def pair_pleio_summary(spec: dict) -> pd.DataFrame:
    path = BASE / f"06_pleiofdr_{spec['pair']}" / f"pleiofdr_{spec['pair']}_summary.csv"
    return with_analysis(read_csv(path), spec)


def pair_pleio_loci(spec: dict) -> pd.DataFrame:
    path = BASE / f"06_pleiofdr_{spec['pair']}" / f"{spec['pair']}_conjfdr_0.05_loci.csv"
    return with_analysis(read_csv(path), spec)


def pair_coloc(spec: dict) -> pd.DataFrame:
    path = BASE / f"09_coloc_{spec['pair']}" / f"coloc_{spec['pair']}_regions.tsv"
    return with_analysis(read_tsv(path), spec)


def pair_pwcoco(spec: dict) -> pd.DataFrame:
    priority = BASE / f"10_pwcoco_{spec['pair']}" / f"coloc_pwcoco_{spec['pair']}_priority_regions.tsv"
    fallback = BASE / f"10_pwcoco_{spec['pair']}" / f"pwcoco_{spec['pair']}_best_h4.tsv"
    path = priority if priority.exists() else fallback
    return with_analysis(read_tsv(path), spec)


def locate_fuma_genes(spec: dict) -> Path:
    root = BASE / f"13_fuma_{spec['pair']}"
    candidates = [p for p in root.rglob("genes.txt") if "FUMA_job" in str(p.parent)]
    if not candidates:
        candidates = list(root.rglob("genes.txt"))
    if not candidates:
        raise FileNotFoundError(f"No genes.txt found under {root}")
    candidates.sort(key=lambda p: (len(p.parts), str(p)))
    return candidates[0]


def pair_fuma(spec: dict) -> pd.DataFrame:
    return with_analysis(read_tsv(locate_fuma_genes(spec)), spec)


def pair_snp_evidence(spec: dict) -> pd.DataFrame:
    path = BASE / f"20_snp_evidence_{spec['pair']}" / f"{spec['pair']}_lead_snp_evidence_table.tsv"
    return with_analysis(read_tsv(path), spec)


def pair_ctwas_counts(spec: dict) -> pd.DataFrame:
    parts = []
    for trait in spec["traits"]:
        path = BASE / f"16_ctwas_{spec['pair']}" / "summary" / f"{trait}_counts.tsv"
        df = read_tsv(path)
        parts.append(df)
    return with_analysis(pd.concat(parts, ignore_index=True), spec)


def pair_candidate_shortlist(spec: dict) -> pd.DataFrame:
    path = BASE / f"17_candidate_gene_integration_{spec['pair']}" / f"{spec['pair']}_cross_evidence_shortlist.tsv"
    return with_analysis(read_tsv(path), spec)


def pair_candidate_tiers(spec: dict) -> pd.DataFrame:
    path = BASE / f"17_candidate_gene_integration_{spec['pair']}" / f"{spec['pair']}_cross_evidence_tier_counts.tsv"
    return with_analysis(read_tsv(path), spec)


def pair_brainmeta_summary(spec: dict) -> pd.DataFrame:
    df = read_tsv(BASE / f"11_smr_{spec['pair']}" / "summary" / f"smr_brainmeta_{spec['pair']}_combined.tsv")
    rows = []
    for trait, sub in df.groupby("trait"):
        sub = sub.copy()
        sub["p_SMR_num"] = pd.to_numeric(sub["p_SMR"], errors="coerce")
        sub["p_HEIDI_num"] = pd.to_numeric(sub["p_HEIDI"], errors="coerce")
        best = sub.sort_values("p_SMR_num").iloc[0]
        rows.append(
            {
                "trait": trait,
                "panel": "BrainMeta",
                "context": "bulk_brain",
                "n_rows": len(sub),
                "n_hits_p_SMR_lt_1e4": int((sub["p_SMR_num"] < 1e-4).sum()),
                "n_hits_supported": int(((sub["p_SMR_num"] < 1e-4) & (sub["p_HEIDI_num"] > 0.01)).sum()),
                "best_gene": best["Gene"],
                "best_probe": best.get("probeID", ""),
                "best_topSNP": best["topSNP"],
                "best_p_SMR": best["p_SMR"],
                "best_p_HEIDI": best["p_HEIDI"],
            }
        )
    return with_analysis(pd.DataFrame(rows), spec)


def pair_gtex_counts(spec: dict) -> pd.DataFrame:
    path = BASE / f"14_smr_gtex_{spec['pair']}" / "summary" / f"smr_gtex_{spec['pair']}_counts.tsv"
    return with_analysis(read_tsv(path), spec)


def pair_bryois_counts(spec: dict) -> pd.DataFrame:
    path = BASE / f"15_smr_bryois_{spec['pair']}" / "summary" / f"smr_bryois_{spec['pair']}_counts.tsv"
    return with_analysis(read_tsv(path), spec)


def build_triplet_audit() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "section": "Upstream factor construction",
                "sheet_scope": "S1-S12",
                "status": "F2-only retained",
                "notes": "These sheets describe the original F2 factor construction/Q_SNP/LOO pipeline and are kept unchanged as the project template.",
            },
            {
                "section": "Cross-disease comparative core analyses",
                "sheet_scope": "S3, S13-S25",
                "status": "Triplet-integrated",
                "notes": "F2-AD, nonlipid8_F1-PD, and lipid8_F1-PD are merged on the same method sheets for direct comparison.",
            },
            {
                "section": "FUMA pair mapping",
                "sheet_scope": "S18",
                "status": "Triplet-integrated",
                "notes": "Pair-specific FUMA outputs were matched by folder name and rechecked before merge.",
            },
            {
                "section": "Lead SNP evidence",
                "sheet_scope": "S19",
                "status": "Triplet-integrated",
                "notes": "All three analyses now have lead-SNP evidence tables with conjFDR, coloc, PWCoCo, and FUMA annotations.",
            },
            {
                "section": "Candidate genes",
                "sheet_scope": "S21-S22",
                "status": "Triplet-integrated",
                "notes": "The same cross-evidence logic and columns were used across the three analyses.",
            },
            {
                "section": "Single-cell / KNK",
                "sheet_scope": "S26-S38",
                "status": "F2-only currently",
                "notes": "PD scPagwas2 and downstream KNK are not yet added; no analytical error detected, just pending work.",
            },
        ]
    )


def build_contents(sheetnames: list[str]) -> pd.DataFrame:
    descriptions = {
        "Contents": "Workbook index for the triplet-comparison supplementary tables.",
        "Audit_triplet": "Audit sheet summarizing what has been integrated across the three analyses and what remains F2-only.",
        "S3_rg_panel": "Requested LDSC cross-trait panel rows for the three focal factor-disease pairs.",
        "S13_CrossDisease_LDSC_MiXeR": "Merged LDSC and MiXeR summaries across the three focal pairs.",
        "S14_PleioFDR_summary": "pleioFDR one-line summaries across the three focal pairs.",
        "S15_PleioFDR_loci": "conjFDR loci across the three focal pairs.",
        "S16_Coloc_summary": "Regional coloc summaries across the three focal pairs.",
        "S17_PWCoCo_best": "PWCoCo priority/shared-signal regions across the three focal pairs.",
        "S18_FUMA_mapping_summary": "FUMA gene-mapping results across the three focal pairs.",
        "S19_SNP_evidence": "Lead-SNP evidence tables across the three focal pairs.",
        "S20_cTWAS_overview": "Trait-level cTWAS count summaries across the three focal pairs.",
        "S21_Candidate_master": "Cross-evidence candidate shortlist across the three focal pairs.",
        "S22_Candidate_tier_summary": "Priority tier counts across the three focal pairs.",
        "S23_BulkBrain_SMR_summary": "BrainMeta SMR summaries across the three focal pairs.",
        "S24_GTEx_SMR_summary": "GTEx SMR count summaries across the three focal pairs.",
        "S25_CellType_SMR_summary": "Bryois cell-type SMR count summaries across the three focal pairs.",
    }
    rows = []
    for idx, name in enumerate(sheetnames, start=1):
        rows.append(
            {
                "Supplementary Table": f"S{idx}" if name.startswith("S") else "",
                "Sheet name": name,
                "Description": descriptions.get(name, "Retained from the original F2-AD supplementary workbook."),
            }
        )
    return pd.DataFrame(rows)


def concat_triplet(func) -> pd.DataFrame:
    return pd.concat([func(spec) for spec in PAIR_SPECS], ignore_index=True)


def replace_sheet(wb, name: str):
    idx = wb.sheetnames.index(name)
    ws_old = wb[name]
    wb.remove(ws_old)
    return wb.create_sheet(name, idx)


def main() -> None:
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(TEMPLATE, OUT_XLSX)
    wb = load_workbook(OUT_XLSX)

    replacements = {
        "S3_rg_panel": (
            "Cross-disease rg panel",
            "Three focal factor-disease pairs are shown together for direct comparison.",
            concat_triplet(pair_rg_panel)[["analysis", "pair", "factor_trait", "disease_trait", "trait1", "trait2", "covariance", "covariance_se", "rg", "rg_se", "intercept", "z_cov", "p_cov", "z_rg", "p_rg", "fdr_rg", "fdr_cov"]],
        ),
        "S13_CrossDisease_LDSC_MiXeR": (
            "Cross-disease LDSC + MiXeR",
            "Merged LDSC pairwise rg and MiXeR summaries across lipid8_F2 × AD, nonlipid8_F1 × PD, and lipid8_F1 × PD.",
            concat_triplet(pair_rg_mixer),
        ),
        "S14_PleioFDR_summary": (
            "pleioFDR summary",
            "One-line pleioFDR/conjFDR summaries across the three focal pairs.",
            concat_triplet(pair_pleio_summary),
        ),
        "S15_PleioFDR_loci": (
            "conjFDR loci",
            "Sentinel loci passing conjFDR < 0.05 across the three focal pairs.",
            concat_triplet(pair_pleio_loci),
        ),
        "S16_Coloc_summary": (
            "coloc regional summary",
            "Regional coloc summaries across the three focal pairs.",
            concat_triplet(pair_coloc),
        ),
        "S17_PWCoCo_best": (
            "PWCoCo prioritized regions",
            "Priority/shared-signal PWCoCo results across the three focal pairs.",
            concat_triplet(pair_pwcoco),
        ),
        "S18_FUMA_mapping_summary": (
            "FUMA mapping summary",
            "FUMA gene mapping outputs across the three focal pairs.",
            concat_triplet(pair_fuma),
        ),
        "S19_SNP_evidence": (
            "Lead SNP evidence",
            "Lead-SNP evidence tables across the three focal pairs, combining conjFDR, coloc, PWCoCo, and FUMA annotations.",
            concat_triplet(pair_snp_evidence),
        ),
        "S20_cTWAS_overview": (
            "cTWAS overview",
            "Trait-level cTWAS summary counts across the three focal pairs.",
            concat_triplet(pair_ctwas_counts),
        ),
        "S21_Candidate_master": (
            "Candidate shortlist",
            "High-priority cross-evidence candidate shortlist across the three focal pairs, using the same logic as the original F2-AD workbook.",
            concat_triplet(pair_candidate_shortlist),
        ),
        "S22_Candidate_tier_summary": (
            "Candidate tier summary",
            "Priority tier counts across the three focal pairs.",
            concat_triplet(pair_candidate_tiers),
        ),
        "S23_BulkBrain_SMR_summary": (
            "Bulk-brain SMR summary",
            "BrainMeta SMR summaries across the three focal pairs.",
            concat_triplet(pair_brainmeta_summary),
        ),
        "S24_GTEx_SMR_summary": (
            "GTEx SMR summary",
            "GTEx tissue-level SMR count summaries across the three focal pairs.",
            concat_triplet(pair_gtex_counts),
        ),
        "S25_CellType_SMR_summary": (
            "Cell-type SMR summary",
            "Bryois cell-type SMR count summaries across the three focal pairs.",
            concat_triplet(pair_bryois_counts),
        ),
    }

    for name, (title, subtitle, df) in replacements.items():
        ws = replace_sheet(wb, name)
        write_sheet(ws, title, subtitle, df)

    if "Audit_triplet" in wb.sheetnames:
        wb.remove(wb["Audit_triplet"])
    wb.create_sheet("Audit_triplet", 1)
    write_sheet(
        wb["Audit_triplet"],
        "Triplet integration audit",
        "Audit of what has been integrated across lipid8_F2 × AD, nonlipid8_F1 × PD, and lipid8_F1 × PD, and what remains intentionally F2-only.",
        build_triplet_audit(),
    )

    ws = wb["Contents"]
    write_sheet(
        ws,
        "Contents",
        "Triplet-comparison supplementary workbook: original F2-AD backbone retained, comparable method sheets merged across the three focal analyses.",
        build_contents(wb.sheetnames),
    )

    wb.save(OUT_XLSX)


if __name__ == "__main__":
    main()
